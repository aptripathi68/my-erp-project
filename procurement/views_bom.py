from __future__ import annotations

import tempfile
from io import BytesIO
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_CEILING
import json
import os
import re
import urllib.error
import urllib.request

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

import openpyxl
from openpyxl.utils import get_column_letter

from estimation.models import EstimateProject
from drawings.models import Drawing
from .models import BOMColumnMapping, BOMComponent, BOMHeader, BOMMark
from .services.backbone import duplicate_work_order_exists, normalize_wo_number, sync_bom_to_backbone
from .services.planning import bom_material_evaluation, bom_planning_summary, generate_int_erc_jobs
from .services.bom_importer import build_header_signature, validate_and_extract_workbook, workbook_sheet_headers


MAPPING_FIELDS = (
    "item_description",
    "grade",
    "mark_no",
    "erc_quantity",
    "item_no",
    "qty_all",
    "length",
    "width",
    "unit_wt",
)

STANDARD_BOM_COLUMNS = [
    "Dispatch MKD No",
    "Assembly Qty",
    "Part Mark / Item No",
    "Section / Profile",
    "Grade / Quality",
    "Part-Qty per Assembly",
    "Length mm",
    "Width mm",
    "Weight per Assembly",
    "Remarks",
]

EXTRACTOR_FIELD_LABELS = {
    "dispatch_mkd_no": "Dispatch MKD No",
    "assembly_qty": "Assembly Qty",
    "part_mark": "Part Mark / Item No",
    "section_profile": "Section / Profile",
    "grade_quality": "Grade / Quality",
    "part_qty_per_assembly": "Part-Qty per Assembly",
    "length_mm": "Length mm",
    "width_mm": "Width mm",
    "weight_per_assembly": "Weight per Assembly",
    "remarks": "Remarks",
}

EXTRACTOR_ALIASES = {
    "dispatch_mkd_no": [
        "dispatch mkd no", "dispatch mkg no", "dispatch mark no", "dispatch mkd", "as marked",
        "mark no", "mark", "du material", "material", "bom part no.",
    ],
    "assembly_qty": [
        "assembly qty", "assy qty", "dispatch qty", "du quantity", "mark qty", "qty requested",
    ],
    "part_mark": [
        "part mark", "part mark / item no", "item no", "item", "item code", "component", "material",
    ],
    "section_profile": [
        "section / profile", "section profile", "section", "profile", "material section",
        "material description", "component description", "du material description",
    ],
    "grade_quality": [
        "grade / quality", "grade quality", "grade", "quality", "material grade", "mat grade",
    ],
    "part_qty_per_assembly": [
        "part-qty per assembly", "part qty per assembly", "qty/assy", "qty per assy",
        "qty per assembly", "unit qty", "qty", "qty all", "quantity",
    ],
    "length_mm": ["length mm", "length (mm)", "length", "len"],
    "width_mm": ["width mm", "width (mm)", "width", "wd"],
    "weight_per_assembly": [
        "weight per assembly", "weight/assy.", "weight/assy", "engg. weight", "engg weight",
        "unit wt. (kgs)", "unit wt", "unit weight", "drg wt. (kgs)", "total wt", "total weight",
    ],
    "remarks": ["remarks", "remark", "revision"],
}

EXTRACTOR_TARGET_ALIASES = {
    field: [label.lower(), field.replace("_", " ")]
    for field, label in EXTRACTOR_FIELD_LABELS.items()
}


def _has_planning_access(user):
    return user.is_authenticated and (
        user.is_superuser
        or user.is_staff
        or user.role in {"Admin", "Planning", "Management", "Store", "Procurement"}
    )


def _session_safe_errors(errors):
    safe = []
    for err in errors:
        row = {}
        for k, v in err.items():
            if isinstance(v, list):
                row[k] = [str(x) for x in v]
            elif v is None:
                row[k] = ""
            else:
                row[k] = str(v)
        safe.append(row)
    return safe


def _date_cell_to_iso(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    try:
        return date.fromisoformat(value[:10]).isoformat()
    except ValueError:
        return value


def _parse_decimal(value: str):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return Decimal(value)
    except (InvalidOperation, ValueError):
        return None


def _extractor_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().replace("\n", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _extractor_find_header_row(ws, max_scan_rows: int = 50):
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        headers = [_extractor_header(value) for value in row]
        matches = 0
        for aliases in EXTRACTOR_ALIASES.values():
            normalized_aliases = {_extractor_header(alias) for alias in aliases}
            if any(header in normalized_aliases for header in headers):
                matches += 1
        if matches >= 3:
            return row_number, headers
    return None, []


def _extractor_auto_mapping(headers):
    mapping = {}
    for field, aliases in EXTRACTOR_ALIASES.items():
        for alias in aliases:
            alias_norm = _extractor_header(alias)
            for index, header in enumerate(headers):
                if header and header == alias_norm:
                    mapping[field] = index
                    break
            if field in mapping:
                break
    return mapping


def _extractor_field_from_instruction(target: str):
    target_norm = _extractor_header(target).replace("-", " ")
    for field, aliases in EXTRACTOR_TARGET_ALIASES.items():
        for alias in aliases:
            alias_norm = alias.replace("-", " ")
            if target_norm == alias_norm or target_norm in alias_norm or alias_norm in target_norm:
                return field
    return None


def _extractor_instruction_mapping(headers, instructions: str):
    mapping = {}
    normalized_headers = [_extractor_header(header) for header in headers]
    for line in (instructions or "").splitlines():
        column_match = re.search(
            r"(.+?)\s+(?:is|in|from|=|:).*?\bcolumn\s+([A-Z]+|\d+)\b",
            line,
            flags=re.IGNORECASE,
        )
        if column_match:
            field = _extractor_field_from_instruction(column_match.group(1))
            raw_column = column_match.group(2).strip().upper()
            if field:
                if raw_column.isdigit():
                    mapping[field] = max(0, int(raw_column) - 1)
                else:
                    column_index = 0
                    for char in raw_column:
                        column_index = column_index * 26 + (ord(char) - ord("A") + 1)
                    mapping[field] = max(0, column_index - 1)
                continue

        match = re.search(r"use\s+(.+?)\s+as\s+(.+)", line, flags=re.IGNORECASE)
        if not match:
            continue
        source_text = _extractor_header(match.group(1))
        field = _extractor_field_from_instruction(match.group(2))
        if not field:
            continue
        for index, header in enumerate(normalized_headers):
            if not header:
                continue
            if source_text == header or source_text in header or header in source_text:
                mapping[field] = index
                break
    return mapping


def _extractor_cell(row, mapping, field):
    index = mapping.get(field)
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


def _extractor_blankish(value: str) -> bool:
    return (value or "").strip() in {"", "-", "--", "0"}


def _extractor_numeric_or_blank(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    return value if _parse_decimal(value.replace(",", "")) is not None else ""


def _extractor_trim(value, limit: int = 120) -> str:
    value = "" if value is None else str(value).strip()
    value = re.sub(r"\s+", " ", value)
    if len(value) > limit:
        return value[: limit - 3] + "..."
    return value


def _extractor_row_quality(rows):
    warnings = []
    required = ["dispatch_mkd_no", "section_profile", "grade_quality", "part_qty_per_assembly", "weight_per_assembly"]
    for field in required:
        missing = sum(1 for row in rows if not row.get(field))
        if missing:
            warnings.append(f"{missing} row(s) missing {EXTRACTOR_FIELD_LABELS[field]}.")

    for row in rows:
        errors = []
        for field in required:
            if not row.get(field):
                errors.append(f"Missing {EXTRACTOR_FIELD_LABELS[field]}")
        for field in ["assembly_qty", "part_qty_per_assembly", "length_mm", "width_mm", "weight_per_assembly"]:
            value = (row.get(field) or "").replace(",", "")
            if value and _parse_decimal(value) is None:
                errors.append(f"Invalid number in {EXTRACTOR_FIELD_LABELS[field]}")
        row["errors"] = errors
    return warnings


def _extractor_workbook_context(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    context = []
    total_rows = 0
    max_total_rows = 160

    for sheet_name in wb.sheetnames:
        if len(context) >= 6 or total_rows >= max_total_rows:
            break
        if any(token in sheet_name.lower() for token in ["summary", "notes", "index", "cover"]):
            continue

        ws = wb[sheet_name]
        header_row, headers = _extractor_find_header_row(ws)
        if not header_row:
            header_row = 1
            headers = []

        raw_header_values = []
        header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), [])
        for index, value in enumerate(header_cells[:40], start=1):
            if value is None or str(value).strip() == "":
                continue
            raw_header_values.append({
                "column": get_column_letter(index),
                "index": index,
                "header": _extractor_trim(value),
                "normalized_header": headers[index - 1] if index <= len(headers) else _extractor_header(value),
            })

        data_rows = []
        for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if total_rows >= max_total_rows or len(data_rows) >= 80:
                break
            cells = {}
            for index, value in enumerate(row[:40], start=1):
                text = _extractor_trim(value)
                if text:
                    cells[get_column_letter(index)] = text
            if not cells:
                continue
            data_rows.append({"row_number": row_number, "cells": cells})
            total_rows += 1

        if raw_header_values or data_rows:
            context.append({
                "sheet_name": sheet_name,
                "header_row": header_row,
                "headers": raw_header_values,
                "sample_rows": data_rows,
            })

    return context


def _extractor_json_from_text(text: str):
    text = (text or "").strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text, flags=re.IGNORECASE).strip()
        text = re.sub(r"```$", "", text).strip()
    return json.loads(text)


def _normalize_ai_extracted_rows(payload):
    rows = []
    for raw in payload.get("rows", []):
        if not isinstance(raw, dict):
            continue
        row = {
            "source_sheet": _extractor_trim(raw.get("source_sheet")),
            "source_row": raw.get("source_row") or "",
            "dispatch_mkd_no": _extractor_trim(raw.get("dispatch_mkd_no")),
            "assembly_qty": _extractor_numeric_or_blank(_extractor_trim(raw.get("assembly_qty"))) or "1",
            "part_mark": _extractor_trim(raw.get("part_mark")),
            "section_profile": _extractor_trim(raw.get("section_profile")),
            "grade_quality": _extractor_trim(raw.get("grade_quality")),
            "part_qty_per_assembly": _extractor_numeric_or_blank(_extractor_trim(raw.get("part_qty_per_assembly"))) or "1",
            "length_mm": _extractor_numeric_or_blank(_extractor_trim(raw.get("length_mm"))),
            "width_mm": _extractor_numeric_or_blank(_extractor_trim(raw.get("width_mm"))),
            "weight_per_assembly": _extractor_numeric_or_blank(_extractor_trim(raw.get("weight_per_assembly"))),
            "remarks": _extractor_trim(raw.get("remarks"), limit=240),
        }
        if row["section_profile"] or row["part_mark"] or row["dispatch_mkd_no"]:
            rows.append(row)
    return rows


def _extract_company_bom_rows_with_ai(xlsx_path: str, instructions: str = ""):
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is not configured.")

    workbook_context = _extractor_workbook_context(xlsx_path)
    if not workbook_context:
        raise RuntimeError("No readable workbook rows were found for AI extraction.")

    system_prompt = (
        "You are a BOM extraction agent for a steel fabrication ERP. "
        "Convert uploaded company BOM rows into Kalpadeep's standard BOM upload rows. "
        "Follow user instructions as corrections to your mapping. "
        "Return only valid JSON. Do not include markdown."
    )
    user_prompt = {
        "task": "Extract standard BOM rows from the workbook context.",
        "standard_fields": {
            "source_sheet": "Original Excel sheet name.",
            "source_row": "Original Excel row number used for this extracted row.",
            "dispatch_mkd_no": "Customer/company Dispatch MKD No or dispatch mark identity.",
            "assembly_qty": "Number of separate fabricated assemblies for this dispatch mark.",
            "part_mark": "Child part mark/item/component code.",
            "section_profile": "Section/Profile/Material description of the child part.",
            "grade_quality": "Material grade/quality.",
            "part_qty_per_assembly": "Child part quantity used in one assembly.",
            "length_mm": "Length in mm if available.",
            "width_mm": "Width in mm if available.",
            "weight_per_assembly": "Weight contribution of this child part for one assembly.",
            "remarks": "Any relevant remarks.",
        },
        "rules": [
            "Use numeric values only for quantity, length, width, and weight fields.",
            "If a numeric field is not available, use an empty string, except assembly_qty and part_qty_per_assembly may default to 1.",
            "Do not create total quantity. Do not multiply weight by assembly quantity.",
            "Skip title, note, subtotal, and total rows.",
            "If user instructions mention a column number or letter for a field, obey that mapping.",
            "Include warnings for uncertain mappings or missing required fields.",
        ],
        "user_instructions": instructions or "",
        "workbook": workbook_context,
        "json_schema": {
            "rows": [
                {
                    "source_sheet": "",
                    "source_row": 0,
                    "dispatch_mkd_no": "",
                    "assembly_qty": "",
                    "part_mark": "",
                    "section_profile": "",
                    "grade_quality": "",
                    "part_qty_per_assembly": "",
                    "length_mm": "",
                    "width_mm": "",
                    "weight_per_assembly": "",
                    "remarks": "",
                }
            ],
            "sheets": [
                {"sheet_name": "", "detected": True, "mapping": {"Dispatch MKD No": "Column I / Dispatch MKD No"}}
            ],
            "warnings": [""],
        },
    }

    request_body = {
        "model": os.environ.get("OPENAI_BOM_MODEL", "gpt-4.1-mini"),
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps(user_prompt, ensure_ascii=True)},
        ],
        "response_format": {"type": "json_object"},
        "temperature": 0,
    }
    request = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=json.dumps(request_body).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=75) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"OpenAI extraction failed: {detail[:500]}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"OpenAI extraction failed: {exc.reason}") from exc

    content = response_payload["choices"][0]["message"]["content"]
    payload = _extractor_json_from_text(content)
    rows = _normalize_ai_extracted_rows(payload)
    warnings = [str(w) for w in payload.get("warnings", []) if str(w).strip()]
    warnings.extend(_extractor_row_quality(rows))
    return {
        "rows": rows,
        "sheets": payload.get("sheets") or [],
        "warnings": warnings,
        "extractor_mode": "OpenAI AI Agent",
    }


def _extract_company_bom_rows(xlsx_path: str, instructions: str = ""):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    rows = []
    sheets = []
    for sheet_name in wb.sheetnames:
        if any(token in sheet_name.lower() for token in ["summary", "notes", "index", "cover"]):
            continue
        ws = wb[sheet_name]
        header_row, headers = _extractor_find_header_row(ws)
        if not header_row:
            sheets.append({"sheet_name": sheet_name, "detected": False, "mapping": {}})
            continue

        mapping = _extractor_auto_mapping(headers)
        mapping.update(_extractor_instruction_mapping(headers, instructions))
        sheets.append({
            "sheet_name": sheet_name,
            "detected": True,
            "header_row": header_row,
            "mapping": {
                EXTRACTOR_FIELD_LABELS[field]: headers[index]
                for field, index in mapping.items()
                if index < len(headers)
            },
        })

        for excel_row, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not row or all(value is None or str(value).strip() == "" for value in row):
                continue
            assembly_qty = _extractor_numeric_or_blank(_extractor_cell(row, mapping, "assembly_qty")) or "1"
            part_qty = _extractor_numeric_or_blank(_extractor_cell(row, mapping, "part_qty_per_assembly")) or "1"
            length_mm = _extractor_numeric_or_blank(_extractor_cell(row, mapping, "length_mm"))
            width_mm = _extractor_numeric_or_blank(_extractor_cell(row, mapping, "width_mm"))
            weight_per_assembly = _extractor_numeric_or_blank(_extractor_cell(row, mapping, "weight_per_assembly"))
            extracted = {
                "source_sheet": sheet_name,
                "source_row": excel_row,
                "dispatch_mkd_no": _extractor_cell(row, mapping, "dispatch_mkd_no"),
                "assembly_qty": assembly_qty,
                "part_mark": _extractor_cell(row, mapping, "part_mark"),
                "section_profile": _extractor_cell(row, mapping, "section_profile"),
                "grade_quality": _extractor_cell(row, mapping, "grade_quality"),
                "part_qty_per_assembly": part_qty,
                "length_mm": length_mm,
                "width_mm": width_mm,
                "weight_per_assembly": weight_per_assembly,
                "remarks": _extractor_cell(row, mapping, "remarks"),
            }
            if _extractor_blankish(extracted["dispatch_mkd_no"]):
                extracted["dispatch_mkd_no"] = ""
            if (
                _extractor_blankish(extracted["section_profile"])
                and _extractor_blankish(extracted["part_mark"])
            ):
                continue
            rows.append(extracted)
    warnings = _extractor_row_quality(rows)
    return {"rows": rows, "sheets": sheets, "warnings": warnings, "extractor_mode": "Rule fallback"}


def _extractor_mapping_sheets(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheets = []
    for sheet_index, sheet_name in enumerate(wb.sheetnames):
        if any(token in sheet_name.lower() for token in ["summary", "notes", "index", "cover"]):
            continue
        ws = wb[sheet_name]
        header_row, headers = _extractor_find_header_row(ws)
        if not header_row:
            header_row = 1
            raw_header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
            headers = [_extractor_header(value) for value in raw_header_values]
        else:
            raw_header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), [])

        columns = []
        for column_index, value in enumerate(raw_header_values, start=1):
            if value is None or str(value).strip() == "":
                continue
            columns.append({
                "index": column_index - 1,
                "letter": get_column_letter(column_index),
                "header": _extractor_trim(value),
            })

        if not columns:
            continue

        default_mapping = _extractor_auto_mapping(headers)
        default_selected = (
            "bom" in sheet_name.lower()
            or (
                "dispatch_mkd_no" in default_mapping
                and "section_profile" in default_mapping
            )
        )
        sheets.append({
            "index": sheet_index,
            "sheet_name": sheet_name,
            "header_row": header_row,
            "columns": columns,
            "selected": default_selected,
            "default_mapping": {field: str(index) for field, index in default_mapping.items()},
            "field_options": [
                {
                    "field": field,
                    "label": label,
                    "selected": str(default_mapping.get(field, "")),
                }
                for field, label in EXTRACTOR_FIELD_LABELS.items()
            ],
        })
    return sheets


def _mapping_display(sheet_mappings):
    display = []
    for mapping in sheet_mappings:
        display.append({
            "sheet_name": mapping["sheet_name"],
            "detected": True,
            "header_row": mapping["header_row"],
            "mapping": {
                EXTRACTOR_FIELD_LABELS[field]: mapping["column_labels"].get(field, "")
                for field in EXTRACTOR_FIELD_LABELS
                if mapping["mapping"].get(field) is not None
            },
        })
    return display


def _posted_extractor_sheet_mappings(request, mapping_sheets):
    sheet_mappings = []
    for sheet in mapping_sheets:
        if request.POST.get(f"use_sheet__{sheet['index']}") != "1":
            continue
        mapping = {}
        labels = {}
        for field in EXTRACTOR_FIELD_LABELS:
            raw_index = request.POST.get(f"map__{sheet['index']}__{field}", "").strip()
            if raw_index == "":
                continue
            try:
                column_index = int(raw_index)
            except ValueError:
                continue
            mapping[field] = column_index
            column = next((col for col in sheet["columns"] if col["index"] == column_index), None)
            if column:
                labels[field] = f"{column['letter']} - {column['header']}"

        if mapping:
            sheet_mappings.append({
                "sheet_name": sheet["sheet_name"],
                "header_row": sheet["header_row"],
                "mapping": mapping,
                "column_labels": labels,
            })
    return sheet_mappings


def _extract_company_bom_rows_from_mapping(xlsx_path: str, sheet_mappings):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    rows = []
    for sheet_mapping in sheet_mappings:
        sheet_name = sheet_mapping["sheet_name"]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        mapping = sheet_mapping["mapping"]
        header_row = sheet_mapping["header_row"]
        for excel_row, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not row or all(value is None or str(value).strip() == "" for value in row):
                continue
            assembly_qty = _extractor_numeric_or_blank(_extractor_cell(row, mapping, "assembly_qty")) or "1"
            part_qty = _extractor_numeric_or_blank(_extractor_cell(row, mapping, "part_qty_per_assembly")) or "1"
            extracted = {
                "source_sheet": sheet_name,
                "source_row": excel_row,
                "dispatch_mkd_no": _extractor_cell(row, mapping, "dispatch_mkd_no"),
                "assembly_qty": assembly_qty,
                "part_mark": _extractor_cell(row, mapping, "part_mark"),
                "section_profile": _extractor_cell(row, mapping, "section_profile"),
                "grade_quality": _extractor_cell(row, mapping, "grade_quality"),
                "part_qty_per_assembly": part_qty,
                "length_mm": _extractor_numeric_or_blank(_extractor_cell(row, mapping, "length_mm")),
                "width_mm": _extractor_numeric_or_blank(_extractor_cell(row, mapping, "width_mm")),
                "weight_per_assembly": _extractor_numeric_or_blank(_extractor_cell(row, mapping, "weight_per_assembly")),
                "remarks": _extractor_cell(row, mapping, "remarks"),
            }
            if (
                _extractor_blankish(extracted["dispatch_mkd_no"])
                and _extractor_blankish(extracted["section_profile"])
                and _extractor_blankish(extracted["part_mark"])
            ):
                continue
            rows.append(extracted)

    warnings = _extractor_row_quality(rows)
    return {
        "rows": rows,
        "sheets": _mapping_display(sheet_mappings),
        "warnings": warnings,
        "extractor_mode": "Manual column mapping",
    }


def _extract_company_bom_rows_for_upload(xlsx_path: str, instructions: str = ""):
    try:
        return _extract_company_bom_rows_with_ai(xlsx_path, instructions)
    except Exception as exc:
        extraction = _extract_company_bom_rows(xlsx_path, instructions)
        extraction["warnings"] = [
            f"AI agent could not complete extraction, so rule fallback was used. Reason: {exc}",
            *extraction.get("warnings", []),
        ]
        extraction["extractor_mode"] = "Rule fallback"
        return extraction


def _standard_bom_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM_UPLOAD"
    ws.append(STANDARD_BOM_COLUMNS)
    for row in rows:
        ws.append([
            row.get("dispatch_mkd_no", ""),
            row.get("assembly_qty", ""),
            row.get("part_mark", ""),
            row.get("section_profile", ""),
            row.get("grade_quality", ""),
            row.get("part_qty_per_assembly", ""),
            row.get("length_mm", ""),
            row.get("width_mm", ""),
            row.get("weight_per_assembly", ""),
            row.get("remarks", ""),
        ])
    for index, width in enumerate([24, 14, 18, 22, 20, 22, 12, 12, 20, 30], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    return wb


def _save_standard_bom_temp(rows):
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb = _standard_bom_workbook(rows)
    wb.save(tmp.name)
    return tmp.name


def _quantity_to_count(value) -> int:
    qty = Decimal(value or "0")
    if qty <= 0:
        return 0
    return int(qty.to_integral_value(rounding=ROUND_CEILING))


def _mark_token(value: str) -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "-", value or "").strip("-").upper()
    return token or "MARK"


def _internal_erc_mark(wo_number: str, dispatch_mkd_no: str, sequence: int) -> str:
    suffix = f"{_mark_token(dispatch_mkd_no)}-{sequence}"
    prefix_budget = max(1, 99 - len(suffix))
    prefix = _mark_token(wo_number)[:prefix_budget].strip("-") or "WO"
    return f"{prefix}-{suffix}"[:100]


def _create_working_bom(
    *,
    result,
    estimate_project,
    bom_name,
    project_name,
    client_name,
    purchase_order_no,
    purchase_order_date,
    delivery_date,
    order_rate,
    order_value,
    user,
) -> BOMHeader:
    with transaction.atomic():
        header = BOMHeader.objects.create(
            estimate_project=estimate_project,
            bom_name=bom_name,
            project_name=project_name,
            client_name=client_name,
            purchase_order_no=purchase_order_no,
            purchase_order_date=purchase_order_date,
            delivery_date=delivery_date,
            order_rate=order_rate,
            order_value=order_value,
            uploaded_by=user,
            uploaded_at=timezone.now(),
        )

        rows_by_dispatch = {}
        for row in result["extracted"]:
            key = (row.sheet_name, row.mark_no or "")
            rows_by_dispatch.setdefault(key, []).append(row)

        components = []
        for (sheet_name, dispatch_mkd_no), rows in rows_by_dispatch.items():
            assembly_qty = max(_quantity_to_count(row.erc_quantity) for row in rows) or 1
            for sequence in range(1, assembly_qty + 1):
                bom_mark = BOMMark.objects.create(
                    bom=header,
                    sheet_name=sheet_name,
                    erc_mark=_internal_erc_mark(bom_name, dispatch_mkd_no, sequence),
                    erc_quantity=Decimal("1"),
                    main_section=rows[0].item_description_raw or "",
                    drawing_no=dispatch_mkd_no or "",
                    drawing=None,
                )
                for row in rows:
                    components.append(
                        BOMComponent(
                            bom_mark=bom_mark,
                            part_mark=row.item_no or "",
                            section_name=row.item_description_raw or "",
                            grade_name=getattr(row, "grade_raw", "") or "",
                            part_quantity_per_assy=row.qty_all,
                            length_mm=row.length_mm,
                            width_mm=row.width_mm,
                            engg_weight_kg=row.line_weight_kg,
                            item_id=row.item_id,
                            item_description_raw=row.item_description_raw or "",
                            excel_row=row.excel_row,
                        )
                    )

        BOMComponent.objects.bulk_create(components, batch_size=2000)
        return header


def _build_user_sheet_mappings(request, headers_info):
    user_sheet_mappings = {}

    for sheet_name, info in headers_info.items():
        if not info.get("detected"):
            continue

        user_sheet_mappings[sheet_name] = {
            "item_description": request.POST.get(f"{sheet_name}__item_description", "").strip(),
            "grade": request.POST.get(f"{sheet_name}__grade", "").strip(),
            "mark_no": request.POST.get(f"{sheet_name}__mark_no", "").strip(),
            "erc_quantity": request.POST.get(f"{sheet_name}__erc_quantity", "").strip(),
            "item_no": request.POST.get(f"{sheet_name}__item_no", "").strip(),
            "qty_all": request.POST.get(f"{sheet_name}__qty_all", "").strip(),
            "length": request.POST.get(f"{sheet_name}__length", "").strip(),
            "width": request.POST.get(f"{sheet_name}__width", "").strip(),
            "unit_wt": request.POST.get(f"{sheet_name}__unit_wt", "").strip(),
        }

    return user_sheet_mappings


def _has_any_mapping(sheet_mappings):
    if not sheet_mappings:
        return False

    for sheet_map in sheet_mappings.values():
        if any(v for v in sheet_map.values()):
            return True
    return False


def _clean_sheet_mapping(mapping):
    if not isinstance(mapping, dict):
        return {}
    return {
        field: (mapping.get(field) or "").strip()
        for field in MAPPING_FIELDS
    }


def _load_persisted_mappings(headers_info):
    loaded = {}

    for sheet_name, info in headers_info.items():
        if not info.get("detected"):
            continue

        signature = info.get("header_signature") or build_header_signature(info.get("headers", []))
        if not signature:
            continue

        saved = (
            BOMColumnMapping.objects
            .filter(sheet_name=sheet_name, header_signature=signature)
            .order_by("-updated_at")
            .first()
        )
        if saved and saved.mapping:
            loaded[sheet_name] = _clean_sheet_mapping(saved.mapping)

    return loaded


def _persist_sheet_mappings(headers_info, sheet_mappings, user):
    for sheet_name, info in headers_info.items():
        if not info.get("detected"):
            continue

        mapping = _clean_sheet_mapping((sheet_mappings or {}).get(sheet_name, {}))
        if not any(mapping.values()):
            continue

        signature = info.get("header_signature") or build_header_signature(info.get("headers", []))
        if not signature:
            continue

        obj, created = BOMColumnMapping.objects.get_or_create(
            sheet_name=sheet_name,
            header_signature=signature,
            defaults={
                "mapping": mapping,
                "created_by": user,
                "updated_by": user,
            },
        )

        if not created:
            obj.mapping = mapping
            obj.updated_by = user
            obj.save(update_fields=["mapping", "updated_by", "updated_at"])


@login_required
def planning_dashboard(request):
    if not _has_planning_access(request.user):
        messages.error(request, "You do not have permission to access planning and material evaluation.")
        return redirect("dashboard_home")

    return render(
        request,
        "procurement/planning_dashboard.html",
        {
            "summaries": bom_planning_summary(),
        },
    )


@login_required
def planning_bom_detail(request, bom_id: int):
    if not _has_planning_access(request.user):
        messages.error(request, "You do not have permission to access planning and material evaluation.")
        return redirect("dashboard_home")

    bom = get_object_or_404(BOMHeader, pk=bom_id)
    evaluation = bom_material_evaluation(bom)
    jobs = (
        bom.marks.prefetch_related("fabrication_jobs")
        .order_by("sheet_name", "erc_mark")
    )
    return render(
        request,
        "procurement/planning_bom_detail.html",
        {
            "bom": bom,
            "evaluation": evaluation,
            "marks": jobs,
        },
    )


@login_required
def generate_bom_int_erc(request, bom_id: int):
    if request.method != "POST":
        return redirect("procurement:planning_bom_detail", bom_id=bom_id)
    if not _has_planning_access(request.user):
        messages.error(request, "You do not have permission to generate INT-ERC units.")
        return redirect("dashboard_home")

    bom = get_object_or_404(BOMHeader, pk=bom_id)
    result = generate_int_erc_jobs(bom)
    messages.success(
        request,
        (
            "INT-ERC generation completed. "
            f"New units: {result['created_jobs']}, component rows: {result['created_components']}."
        ),
    )
    return redirect("procurement:planning_bom_detail", bom_id=bom.id)


@staff_member_required
def download_standard_bom_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM_UPLOAD"
    ws.append(STANDARD_BOM_COLUMNS)

    instructions = wb.create_sheet("Instructions")
    instructions.append(["Instruction", "Details"])
    instructions.append(["Do not rename BOM_UPLOAD headers", "Paste or enter all BOM rows below the header row in BOM_UPLOAD."])
    instructions.append(["BOM header details", "WO Number, project, client, PO, delivery, rate, and value are entered once on the upload screen."])
    instructions.append(["Dispatch MKD No", "Customer/company identity of the fabricated item."])
    instructions.append(["Assembly Qty", "Number of separate fabricated units for this Dispatch MKD No."])
    instructions.append(["Section / Profile", "Must match Item Master section/profile text."])
    instructions.append(["Grade / Quality", "Must match Item Master grade. Example: IS:2062 E250BR."])
    instructions.append(["Part-Qty per Assembly", "Quantity of this child part in one fabricated unit."])
    instructions.append(["Weight per Assembly", "Weight contribution of this child part for one fabricated unit."])
    instructions.append(["Working BOM", "After validation, Create Working BOM will create one internal ERC/Main Mark for each Assembly Qty."])

    widths = {
        "A": 24, "B": 14, "C": 18, "D": 22, "E": 20, "F": 22,
        "G": 12, "H": 12, "I": 20, "J": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="kalpadeep_standard_bom_template.xlsx"'
    wb.save(resp)
    return resp


@staff_member_required
def ai_bom_extractor(request):
    context = {
        "bom_name": request.POST.get("bom_name") or request.session.get("extractor_bom_name", ""),
        "project_name": request.POST.get("project_name") or request.session.get("extractor_project_name", ""),
        "client_name": request.POST.get("client_name") or request.session.get("extractor_client_name", ""),
        "purchase_order_no": request.POST.get("purchase_order_no") or request.session.get("extractor_purchase_order_no", ""),
        "purchase_order_date": request.POST.get("purchase_order_date") or request.session.get("extractor_purchase_order_date", ""),
        "delivery_date": request.POST.get("delivery_date") or request.session.get("extractor_delivery_date", ""),
        "order_rate": request.POST.get("order_rate") or request.session.get("extractor_order_rate", ""),
        "order_value": request.POST.get("order_value") or request.session.get("extractor_order_value", ""),
        "instructions": request.POST.get("instructions") or request.session.get("extractor_instructions", ""),
    }

    action = request.POST.get("action", "")

    if request.method == "POST":
        for key in [
            "bom_name", "project_name", "client_name", "purchase_order_no",
            "purchase_order_date", "delivery_date", "order_rate", "order_value",
        ]:
            request.session[f"extractor_{key}"] = context[key]

        if request.FILES.get("file"):
            upload = request.FILES["file"]
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                for chunk in upload.chunks():
                    tmp.write(chunk)
                request.session["extractor_tmp_path"] = tmp.name

        tmp_path = request.session.get("extractor_tmp_path")
        if not tmp_path:
            context["error"] = "Upload a company BOM file first."
            return render(request, "procurement/ai_bom_extractor.html", context)

        mapping_sheets = _extractor_mapping_sheets(tmp_path)
        context["mapping_sheets"] = mapping_sheets
        context["field_labels"] = EXTRACTOR_FIELD_LABELS

        if action in {"download_standard", "send_validation"}:
            rows = request.session.get("extractor_rows") or []
            if not rows:
                context["error"] = "Apply column mapping and preview rows before continuing."
                return render(request, "procurement/ai_bom_extractor.html", context)

            if action == "download_standard":
                wb = _standard_bom_workbook(rows)
                resp = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                resp["Content-Disposition"] = 'attachment; filename="mapped_standard_bom.xlsx"'
                wb.save(resp)
                return resp

            standard_path = _save_standard_bom_temp(rows)
            result = validate_and_extract_workbook(standard_path)
            request.session["bom_tmp_path"] = standard_path
            request.session["bom_name"] = context["bom_name"]
            request.session["project_name"] = context["project_name"]
            request.session["client_name"] = context["client_name"]
            request.session["purchase_order_no"] = context["purchase_order_no"]
            request.session["purchase_order_date"] = context["purchase_order_date"]
            request.session["delivery_date"] = context["delivery_date"]
            request.session["order_rate"] = context["order_rate"]
            request.session["order_value"] = context["order_value"]
            request.session["estimate_project_id"] = None
            request.session["bom_validation_errors"] = _session_safe_errors(result.get("errors", []))
            return render(
                request,
                "procurement/bom_upload.html",
                {
                    "result": result,
                    "uploaded_step": True,
                    "bom_name": context["bom_name"],
                    "project_name": context["project_name"],
                    "client_name": context["client_name"],
                    "purchase_order_no": context["purchase_order_no"],
                    "purchase_order_date": context["purchase_order_date"],
                    "delivery_date": context["delivery_date"],
                    "order_rate": context["order_rate"],
                    "order_value": context["order_value"],
                },
            )

        if action in {"extract", ""}:
            if not mapping_sheets:
                context["error"] = "No usable header row was found in the uploaded BOM."
            return render(request, "procurement/ai_bom_extractor.html", context)

        if action == "apply_mapping":
            sheet_mappings = _posted_extractor_sheet_mappings(request, mapping_sheets)
            if not sheet_mappings:
                context["error"] = "Select at least one BOM column mapping."
                return render(request, "procurement/ai_bom_extractor.html", context)
            extraction = _extract_company_bom_rows_from_mapping(tmp_path, sheet_mappings)
        elif action == "reextract":
            request.session["extractor_instructions"] = context["instructions"]
            extraction = _extract_company_bom_rows_for_upload(tmp_path, context["instructions"])
        else:
            extraction = _extract_company_bom_rows_from_mapping(
                tmp_path,
                _posted_extractor_sheet_mappings(request, mapping_sheets),
            )

        request.session["extractor_rows"] = extraction["rows"]
        context.update(extraction)
        context["preview_rows"] = extraction["rows"][:100]

    return render(request, "procurement/ai_bom_extractor.html", context)


@staff_member_required
def bom_upload(request):
    context = {}
    estimate_project = None
    estimate_id = request.GET.get("estimate_project") or request.POST.get("estimate_project")
    if estimate_id:
        try:
            estimate_project = EstimateProject.objects.get(pk=estimate_id)
        except EstimateProject.DoesNotExist:
            estimate_project = None

    # STEP 1: Upload standard template -> validate immediately
    if request.method == "POST" and request.FILES.get("file"):
        f = request.FILES["file"]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            for chunk in f.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        bom_name = (request.POST.get("bom_name") or (estimate_project.work_order_no if estimate_project else "")).strip()
        project_name = (request.POST.get("project_name") or (estimate_project.project_name if estimate_project else "")).strip()
        client_name = (request.POST.get("client_name") or (estimate_project.client_name if estimate_project else "")).strip()
        purchase_order_no = (request.POST.get("purchase_order_no") or (estimate_project.purchase_order_no if estimate_project else "")).strip()
        purchase_order_date = (request.POST.get("purchase_order_date") or (estimate_project.purchase_order_date.isoformat() if estimate_project and estimate_project.purchase_order_date else "")).strip()
        delivery_date = (request.POST.get("delivery_date") or (estimate_project.delivery_date.isoformat() if estimate_project and estimate_project.delivery_date else "")).strip()
        order_rate = (request.POST.get("order_rate") or "").strip()
        order_value = (request.POST.get("order_value") or "").strip()

        result = validate_and_extract_workbook(tmp_path)
        request.session["bom_validation_errors"] = _session_safe_errors(result.get("errors", []))

        request.session["bom_tmp_path"] = tmp_path
        request.session["bom_name"] = bom_name
        request.session["project_name"] = project_name
        request.session["client_name"] = client_name
        request.session["purchase_order_no"] = purchase_order_no
        request.session["purchase_order_date"] = purchase_order_date
        request.session["delivery_date"] = delivery_date
        request.session["order_rate"] = order_rate
        request.session["order_value"] = order_value
        request.session["estimate_project_id"] = estimate_project.id if estimate_project else None

        context["bom_name"] = bom_name
        context["project_name"] = project_name
        context["client_name"] = client_name
        context["purchase_order_no"] = purchase_order_no
        context["purchase_order_date"] = purchase_order_date
        context["delivery_date"] = delivery_date
        context["order_rate"] = order_rate
        context["order_value"] = order_value
        context["estimate_project"] = estimate_project
        context["uploaded_step"] = True
        context["result"] = result

        return render(request, "procurement/bom_upload.html", context)

    # STEP 2: Re-validate / Create working BOM from the uploaded standard template
    if request.method == "POST" and request.POST.get("action") in ["validate", "create_working_bom"]:
        tmp_path = request.session.get("bom_tmp_path")
        bom_name = request.session.get("bom_name", "Uploaded BOM")
        project_name = request.session.get("project_name", "")
        client_name = request.session.get("client_name", "")
        purchase_order_no = request.session.get("purchase_order_no", "")
        purchase_order_date_raw = request.session.get("purchase_order_date", "")
        delivery_date_raw = request.session.get("delivery_date", "")
        order_rate_raw = request.session.get("order_rate", "")
        order_value_raw = request.session.get("order_value", "")
        estimate_project_id = request.session.get("estimate_project_id")

        if not tmp_path:
            context["error"] = "Please upload the BOM file first."
            return render(request, "procurement/bom_upload.html", context)

        estimate_project = None
        if estimate_project_id:
            try:
                estimate_project = EstimateProject.objects.get(pk=estimate_project_id)
            except EstimateProject.DoesNotExist:
                estimate_project = None

        result = validate_and_extract_workbook(tmp_path)

        request.session["bom_validation_errors"] = _session_safe_errors(result.get("errors", []))

        context["result"] = result
        context["bom_name"] = bom_name
        context["project_name"] = project_name
        context["client_name"] = client_name
        context["purchase_order_no"] = purchase_order_no
        context["purchase_order_date"] = purchase_order_date_raw
        context["delivery_date"] = delivery_date_raw
        context["order_rate"] = order_rate_raw
        context["order_value"] = order_value_raw
        context["estimate_project"] = estimate_project
        context["uploaded_step"] = True

        if request.POST.get("action") == "create_working_bom" and result["ok"]:
            wo_number = normalize_wo_number(bom_name)
            if not wo_number:
                context["error"] = "WO Number is required."
                return render(request, "procurement/bom_upload.html", context)
            if duplicate_work_order_exists(wo_number):
                context["error"] = "This WO already exists."
                return render(request, "procurement/bom_upload.html", context)

            purchase_order_date = None
            if purchase_order_date_raw:
                try:
                    purchase_order_date = date.fromisoformat(purchase_order_date_raw)
                except ValueError:
                    purchase_order_date = None

            delivery_date = None
            if delivery_date_raw:
                try:
                    delivery_date = date.fromisoformat(delivery_date_raw)
                except ValueError:
                    delivery_date = None

            order_rate = _parse_decimal(order_rate_raw)
            order_value = _parse_decimal(order_value_raw)

            header = _create_working_bom(
                result=result,
                estimate_project=estimate_project,
                bom_name=bom_name,
                project_name=project_name,
                client_name=client_name,
                purchase_order_no=purchase_order_no,
                purchase_order_date=purchase_order_date,
                delivery_date=delivery_date,
                order_rate=order_rate,
                order_value=order_value,
                user=request.user,
            )
            request.session["working_bom_id"] = header.id
            context["working_bom_id"] = header.id

        return render(request, "procurement/bom_upload.html", context)

    context["selected_mappings"] = request.session.get("bom_selected_mappings", {})
    context["bom_name"] = request.session.get("bom_name", "")
    context["project_name"] = request.session.get("project_name", "")
    context["client_name"] = request.session.get("client_name", "")
    context["purchase_order_no"] = request.session.get("purchase_order_no", "")
    context["purchase_order_date"] = request.session.get("purchase_order_date", "")
    context["delivery_date"] = request.session.get("delivery_date", "")
    context["order_rate"] = request.session.get("order_rate", "")
    context["order_value"] = request.session.get("order_value", "")
    context["working_bom_id"] = request.session.get("working_bom_id")
    context["estimate_project"] = estimate_project
    if estimate_project:
        context["bom_name"] = context["bom_name"] or estimate_project.work_order_no
        context["project_name"] = context["project_name"] or estimate_project.project_name
        context["client_name"] = context["client_name"] or estimate_project.client_name
        context["purchase_order_no"] = context["purchase_order_no"] or estimate_project.purchase_order_no
        context["purchase_order_date"] = (
            context["purchase_order_date"]
            or (estimate_project.purchase_order_date.isoformat() if estimate_project.purchase_order_date else "")
        )
        context["delivery_date"] = (
            context["delivery_date"]
            or (estimate_project.delivery_date.isoformat() if estimate_project.delivery_date else "")
        )
    return render(request, "procurement/bom_upload.html", context)


@staff_member_required
def download_bom_validation_errors(request):
    errors = request.session.get("bom_validation_errors", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM Validation Errors"

    headers = [
        "Error Type",
        "Sheet Name",
        "Excel Row",
        "ERC Mark",
        "Part Mark",
        "Section Name",
        "Grade Name",
        "Engg Weight (kg)",
        "Message",
        "Normalized Section",
        "Normalized Grade",
        "Possible Grades For Section",
        "Suggestions",
    ]
    ws.append(headers)

    for err in errors:
        possible_grades = err.get("possible_grades_for_section", [])
        suggestions = err.get("suggestions", [])

        ws.append([
            err.get("type", ""),
            err.get("sheet_name", ""),
            err.get("excel_row", ""),
            err.get("mark_no", ""),
            err.get("item_no", ""),
            err.get("item_description_in_bom", ""),
            err.get("grade_in_bom", ""),
            err.get("unit_weight_in_bom", ""),
            err.get("message", ""),
            err.get("normalized_section", ""),
            err.get("normalized_grade", ""),
            ", ".join(possible_grades) if isinstance(possible_grades, list) else str(possible_grades or ""),
            ", ".join(suggestions) if isinstance(suggestions, list) else str(suggestions or ""),
        ])

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="bom_validation_errors.xlsx"'
    return response


@staff_member_required
def bom_export_master(request, bom_id: int):
    header = get_object_or_404(BOMHeader, id=bom_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM_MASTER"

    ws.append([
        "WO Number",
        "Project Name",
        "Client Name",
        "Purchase Order No",
        "Purchase Order Date",
        "Delivery Date",
        "Order Rate",
        "Order Value",
        "Source Sheet",
        "Dispatch MKD No",
        "Internal ERC/Main Mark",
        "Part Mark / Item No",
        "Section / Profile",
        "Grade / Quality",
        "Item Master Description",
        "Part-Qty per Assembly",
        "Length mm",
        "Width mm",
        "Weight per Assembly",
        "ERC/Main Mark Weight",
        "Source Excel Row",
    ])

    marks = header.marks.prefetch_related("components", "components__item").all()

    for m in marks:
        mark_weight = sum((c.engg_weight_kg or Decimal("0")) for c in m.components.all())
        for c in m.components.all():
            ws.append([
                header.bom_name,
                header.project_name or "",
                header.client_name or "",
                header.purchase_order_no or "",
                header.purchase_order_date.isoformat() if header.purchase_order_date else "",
                header.delivery_date.isoformat() if getattr(header, "delivery_date", None) else "",
                float(header.order_rate) if getattr(header, "order_rate", None) is not None else "",
                float(header.order_value) if getattr(header, "order_value", None) is not None else "",
                m.sheet_name,
                m.drawing_no or "",
                m.erc_mark or "",
                c.part_mark or "",
                c.section_name or "",
                c.grade_name or "",
                c.item.item_description if c.item_id else "",
                float(c.part_quantity_per_assy),
                float(c.length_mm) if c.length_mm is not None else "",
                float(c.width_mm) if c.width_mm is not None else "",
                float(c.engg_weight_kg) if c.engg_weight_kg is not None else "",
                float(mark_weight),
                c.excel_row,
            ])

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="BOM_MASTER_{header.id}.xlsx"'
    wb.save(resp)
    return resp


@staff_member_required
def working_bom_verified(request, bom_id: int):
    if request.method != "POST":
        return redirect("procurement:bom_upload")

    header = get_object_or_404(BOMHeader, id=bom_id)
    backbone_result = sync_bom_to_backbone(header, created_by=request.user)
    header.is_locked = True
    header.save(update_fields=["is_locked"])
    request.session.pop("working_bom_id", None)
    messages.success(
        request,
        (
            "Working BOM verified. "
            f"{backbone_result['created_ercs']} ERC record(s), "
            f"{backbone_result['created_units']} INT-ERC unit(s), and "
            f"{backbone_result['created_requirements']} requirement line(s) created."
        ),
    )
    return redirect("procurement:planning_bom_detail", bom_id=header.id)


@staff_member_required
def working_bom_not_verified(request, bom_id: int):
    if request.method != "POST":
        return redirect("procurement:bom_upload")

    header = get_object_or_404(BOMHeader, id=bom_id)
    header.delete()
    request.session.pop("working_bom_id", None)
    messages.warning(request, "Working BOM discarded. Please correct and upload the BOM again.")
    return redirect("procurement:bom_upload")


@staff_member_required
def bom_delete(request, bom_id):
    header = get_object_or_404(BOMHeader, id=bom_id)

    if request.method == "POST":
        drawing_count = Drawing.objects.filter(project=header).count()
        with transaction.atomic():
            Drawing.objects.filter(project=header).delete()
            header.delete()
        if drawing_count:
            messages.success(request, f"BOM and {drawing_count} linked drawing record(s) deleted successfully.")
        else:
            messages.success(request, "BOM deleted successfully.")
        return redirect("procurement:planning_dashboard")

    linked_drawing_count = Drawing.objects.filter(project=header).count()
    return render(
        request,
        "procurement/bom_delete_confirm.html",
        {
            "header": header,
            "linked_drawing_count": linked_drawing_count,
        },
    )
