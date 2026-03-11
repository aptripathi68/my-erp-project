from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from masters.models import Item, normalize_item_description


DEFAULT_GRADE_CODE = "IS:2062"
DEFAULT_GRADE_NAME = "E250BR"
DEFAULT_GRADE_DISPLAY = f"{DEFAULT_GRADE_CODE} {DEFAULT_GRADE_NAME}"


def _h(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def _to_decimal(v: Any) -> Optional[Decimal]:
    if v is None:
        return None

    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except Exception:
            return None

    s = str(v).strip()
    if s == "":
        return None

    s = s.replace(",", "")

    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def normalize_grade_name(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = s.replace("\n", " ")
    s = re.sub(r"[^A-Z0-9]+", "", s)
    return s


ALIASES = {
    "item_description": [
        "item description",
        "item desc",
        "item-description",
        "member",
        "profile",
        "material description",
        "material desc",
        "section name",
    ],
    "grade": [
        "grade",
        "material grade",
        "mat grade",
        "steel grade",
        "item grade",
        "grade name",
        "spec",
        "material spec",
    ],
    "mark_no": [
        "mark no",
        "mark",
        "mark_no",
        "mark number",
        "fabrication mark no",
        "fab mark no",
        "member mark",
        "m/c mark",
        "mc mark",
    ],
    "erc_quantity": [
        "erc mark quantity",
        "erc quantity",
        "erc qty",
        "mark quantity",
        "mark qty",
        "assembly quantity",
        "assy qty",
    ],
    "drawing_no": [
        "drawing no",
        "dwg no",
        "drg no",
        "drawing",
        "dwg",
        "drg",
    ],
   
    "item_no": [
        "item no",
        "item number",
        "item_no",
        "itm no",
        "sr no",
        "s.no",
        "sno",
    ],
    "qty_all": [
        "qty all",
        "qty",
        "quantity",
        "unit qty",
        "part qty",
        "item qty",
        "required qty",
        "nos",
        "no.",
        "no's",
        "total qty",
    ],
    "length": [
        "length",
        "len",
        "cut length",
        "lg",
        "l (mm)",
        "length (mm)",
    ],
    "width": [
        "width",
        "w",
        "wd",
        "b (mm)",
        "width (mm)",
    ],
    "unit_wt": [
        "unit wt",
        "unit weight",
        "line wt",
        "line weight",
        "wt",
        "weight",
        "unit-wt",
        "unitweight",
        "engg weight",
        "engg weight (kg)",
    ],
}


IGNORE_SHEET_NAME_CONTAINS = [
    "summary",
    "notes",
    "index",
    "cover",
]


@dataclass
class ExtractedRow:
    sheet_name: str
    excel_row: int
    mark_no: str
    erc_quantity: Decimal
    drawing_no: str
    item_no: str
    item_description_raw: str
    grade_raw: str
    item_id: int
    qty_all: Decimal
    length_mm: Optional[Decimal]
    width_mm: Optional[Decimal]
    line_weight_kg: Optional[Decimal]


def detect_header_row(ws, max_scan_rows: int = 40):
    for r_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        headers = [_h(v) for v in row]

        has_item_desc = any(h in ALIASES["item_description"] for h in headers)
        if not has_item_desc:
            continue

        other_hits = 0
        for key in ("grade", "mark_no", "qty_all", "item_no", "length"):
            if any(h in ALIASES[key] for h in headers):
                other_hits += 1

        if other_hits >= 1:
            return r_index, headers

    return None, None


def build_col_map(headers: List[str]) -> Dict[str, int]:
    col_map: Dict[str, int] = {}

    for idx, htxt in enumerate(headers):
        for canonical, aliases in ALIASES.items():
            if canonical in col_map:
                continue
            if htxt in aliases:
                col_map[canonical] = idx
                break

    return col_map


def build_user_col_map(headers: List[str], user_mapping: Dict[str, str]) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    normalized_headers = [_h(h) for h in headers]

    for canonical, selected_header in user_mapping.items():
        if not selected_header:
            continue

        selected_norm = _h(selected_header)
        for idx, header_norm in enumerate(normalized_headers):
            if header_norm == selected_norm:
                col_map[canonical] = idx
                break

    return col_map


def get_cell(row: Tuple[Any], col_map: Dict[str, int], key: str):
    idx = col_map.get(key)
    if idx is None:
        return None
    if idx >= len(row):
        return None
    return row[idx]


def workbook_sheet_headers(xlsx_path: str, max_scan_rows: int = 40) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(
        xlsx_path,
        data_only=True,
        read_only=True,
    )

    result = {}

    for sheet_name in wb.sheetnames:
        if any(x in sheet_name.lower() for x in IGNORE_SHEET_NAME_CONTAINS):
            continue

        ws = wb[sheet_name]
        header_row, headers = detect_header_row(ws, max_scan_rows=max_scan_rows)

        if not header_row or not headers:
            result[sheet_name] = {
                "detected": False,
                "header_row": None,
                "headers": [],
            }
            continue

        result[sheet_name] = {
            "detected": True,
            "header_row": header_row,
            "headers": headers,
        }

    return result


def validate_and_extract_workbook(
    xlsx_path: str,
    user_sheet_mappings: Optional[Dict[str, Dict[str, str]]] = None,
) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(
        xlsx_path,
        data_only=True,
        read_only=True,
    )

    items = list(
        Item.objects.select_related("grade")
        .filter(is_active=True)
        .only(
            "id",
            "section_name",
            "item_description",
            "grade__code",
            "grade__name",
        )
    )

    item_by_section_grade: Dict[Tuple[str, str], Item] = {}

    for it in items:
        section_norm = normalize_item_description(it.section_name or "")
        if not section_norm or not it.grade:
            continue

        grade_code = it.grade.code or ""
        grade_name = it.grade.name or ""

        grade_code_norm = normalize_grade_name(grade_code)
        grade_name_norm = normalize_grade_name(grade_name)
        grade_combined_norm = normalize_grade_name(f"{grade_code} {grade_name}")

        if grade_code_norm:
            item_by_section_grade[(section_norm, grade_code_norm)] = it
        if grade_name_norm:
            item_by_section_grade[(section_norm, grade_name_norm)] = it
        if grade_combined_norm:
            item_by_section_grade[(section_norm, grade_combined_norm)] = it

    extracted: List[ExtractedRow] = []
    errors: List[Dict[str, Any]] = []
    detected: Dict[str, Any] = {}

    sheets_used = 0

    for sheet_name in wb.sheetnames:
        if any(x in sheet_name.lower() for x in IGNORE_SHEET_NAME_CONTAINS):
            continue

        ws = wb[sheet_name]
        header_row, headers = detect_header_row(ws)

        if not header_row or not headers:
            detected[sheet_name] = {
                "skipped": True,
                "reason": "header_not_detected",
            }
            continue

        sheet_mapping = (user_sheet_mappings or {}).get(sheet_name, {})

        if sheet_mapping:
            col_map = build_user_col_map(headers, sheet_mapping)
        else:
            col_map = build_col_map(headers)

        if "item_description" not in col_map:
            detected[sheet_name] = {
                "skipped": True,
                "reason": "item_description_column_not_found",
            }
            continue

        if "grade" not in col_map:
            detected[sheet_name] = {
                "skipped": True,
                "reason": "grade_column_not_found",
            }
            continue

        sheets_used += 1

        detected[sheet_name] = {
            "skipped": False,
            "header_row": header_row,
            "col_map": col_map,
        }

        last_mark = ""
        last_erc_quantity = Decimal("1")
        last_drawing = ""

        for excel_r, row_vals in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not row_vals:
                continue

            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue

            item_desc = get_cell(row_vals, col_map, "item_description")
            grade_val = get_cell(row_vals, col_map, "grade")

            if item_desc is None:
                continue

            item_desc_raw = str(item_desc).strip()
            grade_raw = str(grade_val).strip() if grade_val is not None else ""

            if not item_desc_raw:
                continue

            if not grade_raw or grade_raw.strip() == "":
                grade_raw = DEFAULT_GRADE_DISPLAY

            mark_no = get_cell(row_vals, col_map, "mark_no")
            erc_quantity = get_cell(row_vals, col_map, "erc_quantity")
            drawing_no = get_cell(row_vals, col_map, "drawing_no")
            item_no = get_cell(row_vals, col_map, "item_no")
            qty_all = get_cell(row_vals, col_map, "qty_all")
            length_mm = get_cell(row_vals, col_map, "length")
            width_mm = get_cell(row_vals, col_map, "width")
            unit_wt = get_cell(row_vals, col_map, "unit_wt")

            if mark_no and str(mark_no).strip():
                last_mark = str(mark_no).strip()

            erc_qty_dec = _to_decimal(erc_quantity)
            if erc_qty_dec is not None and erc_qty_dec > 0:
                last_erc_quantity = erc_qty_dec

            if drawing_no and str(drawing_no).strip():
                last_drawing = str(drawing_no).strip()

            item_no_final = str(item_no).strip() if item_no else ""
            qty_dec = _to_decimal(qty_all) or Decimal("1")

            section_norm = normalize_item_description(item_desc_raw)
            

            it = item_by_section_grade.get((section_norm, grade_norm))

            if not it:
                section_matches = [
                    x for x in items
                    if normalize_item_description(x.section_name or "") == section_norm
                ]

                if section_matches:
                    possible_grades = sorted({
                        f"{x.grade.code}, {x.grade.name}".strip(", ")
                        for x in section_matches
                        if x.grade and (x.grade.code or x.grade.name)
                    })
                    errors.append({
                        "type": "GRADE_MISMATCH",
                        "sheet_name": sheet_name,
                        "excel_row": excel_r,
                        "mark_no": last_mark,
                        "item_no": item_no_final,
                        "item_description_in_bom": item_desc_raw,
                        "grade_in_bom": grade_raw,
                        "unit_weight_in_bom": _to_decimal(unit_wt),
                        "normalized_section": section_norm,
                        "normalized_grade": grade_norm,
                        "message": "Section matched in Item Master, but Grade did not match.",
                        "possible_grades_for_section": possible_grades,
                    })
                else:
                    hint = item_desc_raw.replace(" ", "")[:6].lower()

                    sugg = []
                    for x in items:
                        section_name = x.section_name or ""
                        if section_name and hint in normalize_item_description(section_name):
                            grade_text = ""
                            if x.grade:
                                grade_text = f"{x.grade.code}, {x.grade.name}".strip(", ")
                            sugg.append(f"{section_name} | {grade_text}")

                    errors.append({
                        "type": "ITEM_GRADE_MISMATCH",
                        "sheet_name": sheet_name,
                        "excel_row": excel_r,
                        "mark_no": last_mark,
                        "item_no": item_no_final,
                        "item_description_in_bom": item_desc_raw,
                        "grade_in_bom": grade_raw,
                        "unit_weight_in_bom": _to_decimal(unit_wt),
                        "normalized_section": section_norm,
                        "normalized_grade": grade_norm,
                        "message": "No Item Master row found matching both Section Name and Grade.",
                        "suggestions": sugg[:8],
                    })
                continue

            extracted.append(
                ExtractedRow(
                    sheet_name=sheet_name,
                    excel_row=excel_r,
                    mark_no=last_mark,
                    erc_quantity=last_erc_quantity,
                    drawing_no=last_drawing,
                    item_no=item_no_final,
                    item_description_raw=item_desc_raw,
                    grade_raw=grade_raw,
                    item_id=it.id,
                    qty_all=qty_dec,
                    length_mm=_to_decimal(length_mm),
                    width_mm=_to_decimal(width_mm),
                    line_weight_kg=_to_decimal(unit_wt),
                )
            )

    summary = {
        "sheets_total": len(wb.sheetnames),
        "sheets_used": sheets_used,
        "rows_extracted": len(extracted),
        "marks_found": len({
            (x.sheet_name, x.mark_no)
            for x in extracted if x.mark_no
        }),
        "errors": len(errors),
    }

    return {
        "ok": len(errors) == 0,
        "summary": summary,
        "errors": errors,
        "extracted": extracted if len(errors) == 0 else [],
        "detected": detected,
    }