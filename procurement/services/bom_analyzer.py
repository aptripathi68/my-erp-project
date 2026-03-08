from openpyxl import load_workbook

from .bom_mapping import detect_header_row, extract_headers, detect_column_mapping


def analyze_workbook(file):

    wb = load_workbook(file, data_only=True)

    result = {}

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        header_row = detect_header_row(ws)

        if not header_row:
            continue

        headers = extract_headers(ws, header_row)

        auto_mapping = detect_column_mapping(headers)

        result[sheet_name] = {
            "header_row": header_row,
            "headers": headers,
            "mapping": auto_mapping
        }

    return result

