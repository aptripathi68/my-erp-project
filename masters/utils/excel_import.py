# masters/utils/excel_import.py
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from django.db import transaction

from masters.models import Group2, Grade, Item


REQUIRED_COLS = [
    "Item Master ID",
    "Item Description",
    "Grade Name",
    "Group2 Name",
    "Unit Wt. (kg/m)",
]
OPTIONAL_COLS = ["Group1 Name", "Section Name"]


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _to_decimal(v) -> Decimal:
    if v is None or str(v).strip() == "":
        return Decimal("0")
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _mk_code(name: str, fallback: str) -> str:
    # Make a safe code (max 50 chars), deterministic
    base = (name or fallback or "UNK").strip()
    return base[:50]


@transaction.atomic
def import_item_master_xlsx(path: str, batch_id: str = "initial") -> dict:
    wb = load_workbook(path)
    ws = wb.active

    headers = [_norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: headers.index(h) for h in headers if h}

    missing = [c for c in REQUIRED_COLS if c not in col]
    if missing:
        raise ValueError(f"Missing columns in Excel: {missing}")

    created, updated, skipped = 0, 0, 0

    for r in ws.iter_rows(min_row=2):
        item_master_id = _norm(r[col["Item Master ID"]].value)
        if not item_master_id:
            skipped += 1
            continue

        item_description = _norm(r[col["Item Description"]].value)
        grade_name = _norm(r[col["Grade Name"]].value)
        group2_name = _norm(r[col["Group2 Name"]].value)
        unit_weight = _to_decimal(r[col["Unit Wt. (kg/m)"]].value)

        group1_name = _norm(r[col["Group1 Name"]].value) if "Group1 Name" in col else ""
        section_name = _norm(r[col["Section Name"]].value) if "Section Name" in col else ""

        # Group2
        g2_code = _mk_code(group2_name, item_master_id)
        group2, _ = Group2.objects.get_or_create(
            code=g2_code,
            defaults={"name": group2_name or g2_code, "description": ""},
        )
        # keep name updated if Excel has a better value
        if group2_name and group2.name != group2_name:
            group2.name = group2_name
            group2.save(update_fields=["name"])

        # Grade (under Group2)
        grade_code = _mk_code(grade_name, item_master_id)
        grade, _ = Grade.objects.get_or_create(
            group2=group2,
            code=grade_code,
            defaults={"name": grade_name or grade_code, "description": ""},
        )
        if grade_name and grade.name != grade_name:
            grade.name = grade_name
            grade.save(update_fields=["name"])

        defaults = {
            "group2": group2,
            "grade": grade,
            "item_description": item_description,
            "group1_name": group1_name,
            "section_name": section_name,
            "unit_weight": unit_weight,
            "import_batch_id": batch_id,
            "is_active": True,
        }

        obj, was_created = Item.objects.update_or_create(
            item_master_id=item_master_id,
            defaults=defaults,
        )

        if was_created:
            created += 1
        else:
            updated += 1

    return {"created": created, "updated": updated, "skipped": skipped}