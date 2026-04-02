from __future__ import annotations

import os
import tempfile
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.utils import timezone
from django.conf import settings

from masters.models import Item
from masters.models import Group2
from procurement.models import BOMColumnMapping

from .models import (
    EstimateExpense,
    EstimateProject,
    EstimateProjectSupplier,
    EstimateRawMaterialLine,
    EstimateRawMaterialRate,
    EstimateSupplier,
    EstimateSupplierQuotationFile,
)
from .services import (
    PAINT_COMPONENT_CODES,
    create_default_suppliers_if_missing,
    ensure_project_cost_heads,
    DEFAULT_COST_HEAD_CONFIG,
    generate_budget_heads,
    recalculate_cost_heads,
    refresh_budget_totals,
    sync_project_supplier_rates,
)
from .storage import (
    build_supplier_quotation_object_key,
    generate_supplier_quotation_download_url,
    upload_supplier_quotation_file,
)
from .services_tentative_bom import (
    build_header_signature as build_tentative_header_signature,
    build_user_col_map as build_tentative_user_col_map,
    validate_and_extract_tentative_bom,
    workbook_sheet_headers as tentative_bom_sheet_headers,
)


TENTATIVE_BOM_MAPPING_FIELDS = ("section_name", "grade", "gross_weight")


def _format_percentage_for_display(value):
    if value is None:
        return ""
    pct = value * Decimal("100")
    if pct == pct.quantize(Decimal("0.01")):
        return f"{pct.quantize(Decimal('0.01'))}"
    return f"{pct.quantize(Decimal('0.001'))}"


def _format_consumption_for_display(value):
    if value is None:
        return ""
    if value == value.quantize(Decimal("0.01")):
        return f"{value.quantize(Decimal('0.01'))}"
    return f"{value.quantize(Decimal('0.001'))}"


def _parse_decimal(value: str, default: Decimal = Decimal("0")) -> Decimal:
    value = (value or "").strip()
    if not value:
        return default
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError):
        return default


def _parse_date(value: str):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return timezone.datetime.fromisoformat(value).date()
    except ValueError:
        return None


def _can_manage_rates(user) -> bool:
    return user.role in {"Admin", "Marketing", "Management"}


def _can_create_estimate(user) -> bool:
    return user.role in {"Admin", "Planning", "Management"}


def _can_manage_raw_materials(user) -> bool:
    return user.role in {"Admin", "Planning", "Management"}


def _can_manage_costs(user) -> bool:
    return user.role in {"Admin", "Planning", "Management"}


def _can_manage_accounts(user) -> bool:
    return user.role in {"Admin", "Accounts", "Management"}


def _can_manage_decision(user) -> bool:
    return user.role in {"Admin", "Management"}


def _can_delete_estimate(user) -> bool:
    return user.role in {"Admin", "Management"}


def _default_active_sheet(user) -> str:
    if user.role == "Marketing":
        return "rate-finalisation"
    if user.role == "Accounts":
        return "budget-monitoring"
    return "raw-material-selection"


def _project_financial_year_label(project) -> str:
    return project.financial_year_label


def _session_safe_errors(errors):
    safe = []
    for err in errors:
        row = {}
        for key, value in err.items():
            if isinstance(value, list):
                row[key] = [str(x) for x in value]
            elif value is None:
                row[key] = ""
            else:
                row[key] = str(value)
        safe.append(row)
    return safe


def _tentative_key(project_id: int, suffix: str) -> str:
    return f"estimation_tentative_bom_{project_id}_{suffix}"


def _clean_sheet_mapping(mapping):
    if not isinstance(mapping, dict):
        return {}
    return {field: (mapping.get(field) or "").strip() for field in TENTATIVE_BOM_MAPPING_FIELDS}


def _load_persisted_tentative_mappings(headers_info):
    loaded = {}
    for sheet_name, info in headers_info.items():
        if not info.get("detected"):
            continue
        signature = info.get("header_signature") or build_tentative_header_signature(info.get("headers", []))
        if not signature:
            continue
        saved = (
            BOMColumnMapping.objects.filter(sheet_name=sheet_name, header_signature=signature)
            .order_by("-updated_at")
            .first()
        )
        if saved and saved.mapping:
            loaded[sheet_name] = _clean_sheet_mapping(saved.mapping)
    return loaded


def _persist_tentative_mappings(headers_info, sheet_mappings, user):
    for sheet_name, info in headers_info.items():
        if not info.get("detected"):
            continue
        mapping = _clean_sheet_mapping((sheet_mappings or {}).get(sheet_name, {}))
        if not any(mapping.values()):
            continue
        signature = info.get("header_signature") or build_tentative_header_signature(info.get("headers", []))
        if not signature:
            continue
        obj, created = BOMColumnMapping.objects.get_or_create(
            sheet_name=sheet_name,
            header_signature=signature,
            defaults={"mapping": mapping, "created_by": user, "updated_by": user},
        )
        if not created:
            obj.mapping = mapping
            obj.updated_by = user
            obj.save(update_fields=["mapping", "updated_by", "updated_at"])


def _build_tentative_user_sheet_mappings(request, headers_info):
    user_sheet_mappings = {}
    for sheet_name, info in headers_info.items():
        if not info.get("detected"):
            continue
        user_sheet_mappings[sheet_name] = {
            field: request.POST.get(f"{sheet_name}__{field}", "").strip()
            for field in TENTATIVE_BOM_MAPPING_FIELDS
        }
    return user_sheet_mappings


def _has_any_mapping(sheet_mappings):
    if not sheet_mappings:
        return False
    return any(any(v for v in sheet_map.values()) for sheet_map in sheet_mappings.values())


def _clear_tentative_bom_session(request, project_id: int):
    for suffix in ("tmp_path", "headers_info", "selected_mappings", "result"):
        request.session.pop(_tentative_key(project_id, suffix), None)


def _build_estimate_detail_context(request, project):
    tentative_headers_info = request.session.get(_tentative_key(project.id, "headers_info")) or {}
    tentative_selected_mappings = request.session.get(_tentative_key(project.id, "selected_mappings")) or {}
    tentative_result = request.session.get(_tentative_key(project.id, "result")) or None
    if tentative_result and tentative_result.get("aggregated_lines") and not tentative_result.get("total_quantity_mt"):
        total_quantity_mt = sum(
            (
                _parse_decimal(row.get("quantity_mt"), default=Decimal("0"))
                for row in tentative_result.get("aggregated_lines", [])
            ),
            Decimal("0"),
        )
        tentative_result["total_quantity_mt"] = str(total_quantity_mt.quantize(Decimal("0.001")))
    preferred_sheet_names = [
        sheet_name
        for sheet_name, info in tentative_headers_info.items()
        if info.get("preferred")
    ]
    if preferred_sheet_names:
        tentative_headers_info = {
            sheet_name: info
            for sheet_name, info in tentative_headers_info.items()
            if sheet_name in preferred_sheet_names
        }
        tentative_selected_mappings = {
            sheet_name: mapping
            for sheet_name, mapping in tentative_selected_mappings.items()
            if sheet_name in preferred_sheet_names
        }

    supplier_links = list(project.project_suppliers.select_related("supplier"))
    supplier_file_rows = []
    for quotation_file in project.supplier_quotation_files.select_related("supplier", "uploaded_by"):
        download_url = None
        try:
            download_url = generate_supplier_quotation_download_url(quotation_file.file_key)
        except Exception:
            download_url = None
        supplier_file_rows.append((quotation_file, download_url))
    rate_matrix = []
    for line in project.raw_material_lines.all():
        supplier_rate_map = {rate.supplier_id: rate for rate in line.supplier_rates.all()}
        row_rates = [supplier_rate_map.get(link.supplier_id) for link in supplier_links]
        rate_matrix.append((line, row_rates))

    return {
        "project": project,
        "group2_list": Group2.objects.order_by("name"),
        "suppliers": EstimateSupplier.objects.filter(is_active=True),
        "supplier_links": supplier_links,
        "supplier_file_rows": supplier_file_rows,
        "rate_matrix": rate_matrix,
        "cost_heads": [
            {
                "obj": head,
                "display_percentage": (
                    _format_consumption_for_display(head.percentage)
                    if head.code in PAINT_COMPONENT_CODES
                    else _format_percentage_for_display(head.percentage)
                ),
                "display_consumption": _format_consumption_for_display(head.percentage)
                if head.code in PAINT_COMPONENT_CODES
                else "",
                "input_label": "Consumption/MT (LTR)"
                if head.code in PAINT_COMPONENT_CODES
                else "Percentage",
                "rate_label": "Rate/LTR" if head.code in PAINT_COMPONENT_CODES else "Cost per Kg",
                "is_paint_component": head.code in PAINT_COMPONENT_CODES,
            }
            for head in project.cost_heads.all()
        ],
        "budget_heads": project.budget_heads.all(),
        "active_sheet": request.GET.get("sheet") or _default_active_sheet(request.user),
        "can_manage_raw_materials": _can_manage_raw_materials(request.user),
        "can_manage_rates": _can_manage_rates(request.user),
        "can_manage_costs": _can_manage_costs(request.user),
        "can_manage_accounts": _can_manage_accounts(request.user),
        "can_manage_decision": _can_manage_decision(request.user),
        "can_delete_estimate": _can_delete_estimate(request.user),
        "can_edit_planning_notes": request.user.role in {"Admin", "Planning", "Management"},
        "can_edit_marketing_notes": request.user.role in {"Admin", "Marketing", "Management"},
        "can_edit_management_notes": request.user.role in {"Admin", "Management"},
        "can_edit_accounts_notes": request.user.role in {"Admin", "Accounts", "Management"},
        "can_edit_any_notes": request.user.role in {
            "Admin",
            "Planning",
            "Marketing",
            "Management",
            "Accounts",
        },
        "management_estimated_price": project.estimated_price_per_kg,
        "can_reopen_quotation": _can_manage_decision(request.user)
        and project.status in {EstimateProject.Status.APPROVED, EstimateProject.Status.REJECTED}
        and not project.budget_heads.exists(),
        "tentative_headers_info": tentative_headers_info,
        "tentative_selected_mappings": tentative_selected_mappings,
        "tentative_result": tentative_result,
    }


@login_required
def estimate_list(request):
    projects = list(EstimateProject.objects.all().prefetch_related("project_suppliers"))
    fy_choices = sorted({_project_financial_year_label(project) for project in projects}, reverse=True)
    selected_fy = (request.GET.get("fy") or "").strip()
    search = (request.GET.get("q") or "").strip()

    if selected_fy:
        projects = [project for project in projects if _project_financial_year_label(project) == selected_fy]

    if search:
        s = search.lower()
        projects = [
            project for project in projects
            if s in project.inquiry_no.lower()
            or s in project.client_name.lower()
            or s in project.project_name.lower()
            or s in (project.work_order_no or "").lower()
            or s in (project.purchase_order_no or "").lower()
        ]

    quotation_projects = [
        project for project in projects
        if project.status in {
            EstimateProject.Status.DRAFT,
            EstimateProject.Status.RATE_FINALIZATION,
            EstimateProject.Status.UNDER_REVIEW,
            EstimateProject.Status.REJECTED,
        }
    ]
    open_budget_projects = [
        project for project in projects
        if project.status in {
            EstimateProject.Status.APPROVED,
            EstimateProject.Status.PO_RECEIVED,
            EstimateProject.Status.IN_EXECUTION,
        }
    ]
    closed_budget_projects = [
        project for project in projects
        if project.status == EstimateProject.Status.CLOSED
    ]

    return render(
        request,
        "estimation/estimate_list.html",
        {
            "quotation_projects": quotation_projects,
            "open_budget_projects": open_budget_projects,
            "closed_budget_projects": closed_budget_projects,
            "fy_choices": fy_choices,
            "selected_fy": selected_fy,
            "search": search,
            "can_create_estimate": _can_create_estimate(request.user),
            "can_delete_estimate": _can_delete_estimate(request.user),
            "can_manage_accounts": _can_manage_accounts(request.user),
        },
    )


@login_required
def estimate_create(request):
    if not _can_create_estimate(request.user):
        messages.error(request, "Only Planning, Management, or Admin can create a new quotation inquiry.")
        return redirect("estimation:estimate_list")

    if request.method == "POST":
        client_name = (request.POST.get("client_name") or "").strip()
        project_name = (request.POST.get("project_name") or "").strip()
        quantity_mt = _parse_decimal(request.POST.get("quantity_mt"), default=Decimal("0"))
        notes = (request.POST.get("notes") or "").strip()

        if not client_name or not project_name:
            messages.error(request, "Client name and project name are required.")
            return render(request, "estimation/estimate_create.html", {})

        project = EstimateProject.objects.create(
            client_name=client_name,
            project_name=project_name,
            quantity_mt=quantity_mt,
            notes=notes,
            created_by=request.user,
            updated_by=request.user,
        )
        create_default_suppliers_if_missing()
        ensure_project_cost_heads(project)
        recalculate_cost_heads(project)
        messages.success(request, f"Estimation inquiry {project.inquiry_no} created.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    return render(request, "estimation/estimate_create.html", {})


@login_required
def estimate_detail(request, project_id: int):
    project = get_object_or_404(
        EstimateProject.objects.prefetch_related(
            "raw_material_lines__item__grade",
            "raw_material_lines__supplier_rates__supplier",
            "project_suppliers__supplier",
            "cost_heads",
            "budget_heads__expenses",
            "boms",
        ),
        pk=project_id,
    )
    ensure_project_cost_heads(project)
    sync_project_supplier_rates(project)
    recalculate_cost_heads(project)
    refresh_budget_totals(project)
    context = _build_estimate_detail_context(request, project)
    return render(request, "estimation/estimate_detail.html", context)


@login_required
def create_supplier(request):
    if request.method != "POST":
        return redirect("estimation:estimate_list")
    if not _can_manage_rates(request.user):
        messages.error(request, "Only Marketing, Management, or Admin can create suppliers.")
        return redirect(request.POST.get("next") or "estimation:estimate_list")

    name = (request.POST.get("name") or "").strip()
    if not name:
        messages.error(request, "Supplier name is required.")
        return redirect(request.POST.get("next") or "estimation:estimate_list")

    supplier, created = EstimateSupplier.objects.get_or_create(name=name)
    if created:
        supplier.contact_person = (request.POST.get("contact_person") or "").strip()
        supplier.phone = (request.POST.get("phone") or "").strip()
        supplier.email = (request.POST.get("email") or "").strip()
        supplier.save()
        messages.success(request, f"Supplier {supplier.name} created.")
    else:
        messages.info(request, f"Supplier {supplier.name} already exists.")
    return redirect(request.POST.get("next") or "estimation:estimate_list")


@login_required
def add_project_supplier(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_rates(request.user):
        messages.error(request, "Only Marketing, Management, or Admin can add suppliers to the rate sheet.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    supplier_id = request.POST.get("supplier_id")
    supplier = get_object_or_404(EstimateSupplier, pk=supplier_id)
    next_order = project.project_suppliers.count() + 1
    EstimateProjectSupplier.objects.get_or_create(
        project=project,
        supplier=supplier,
        defaults={"column_order": next_order},
    )
    sync_project_supplier_rates(project)
    project.status = EstimateProject.Status.RATE_FINALIZATION
    project.updated_by = request.user
    project.save(update_fields=["status", "updated_by", "updated_at"])
    messages.success(request, f"{supplier.name} added to the rate finalization sheet.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def remove_project_supplier(request, project_id: int, supplier_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_rates(request.user):
        messages.error(request, "Only Marketing, Management, or Admin can remove suppliers from the rate sheet.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    link = get_object_or_404(EstimateProjectSupplier, project=project, supplier_id=supplier_id)
    supplier = link.supplier

    EstimateSupplierQuotationFile.objects.filter(project=project, supplier=supplier).delete()
    EstimateRawMaterialRate.objects.filter(line__project=project, supplier=supplier).delete()
    link.delete()

    for index, row in enumerate(project.project_suppliers.order_by("column_order", "supplier__name"), start=1):
        if row.column_order != index:
            row.column_order = index
            row.save(update_fields=["column_order"])

    for line in project.raw_material_lines.prefetch_related("supplier_rates"):
        line.recalculate_from_rates(save=True)

    recalculate_cost_heads(project)
    project.updated_by = request.user
    project.save(update_fields=["updated_by", "updated_at"])
    messages.success(request, f"{supplier.name} removed from this rate sheet.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def import_tentative_bom(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_raw_materials(request.user):
        messages.error(request, "Only Planning, Management, or Admin can upload tentative BOM for raw material selection.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    action = request.POST.get("action")

    if request.FILES.get("file"):
        upload = request.FILES["file"]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            for chunk in upload.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        headers_info = tentative_bom_sheet_headers(tmp_path)
        selected_mappings = {}
        for sheet_name, info in headers_info.items():
            if info.get("detected"):
                selected_mappings[sheet_name] = info.get("mapping", {})

        persisted = _load_persisted_tentative_mappings(headers_info)
        for sheet_name, mapping in persisted.items():
            if any(mapping.values()):
                selected_mappings[sheet_name] = mapping

        request.session[_tentative_key(project.id, "tmp_path")] = tmp_path
        request.session[_tentative_key(project.id, "headers_info")] = headers_info
        request.session[_tentative_key(project.id, "selected_mappings")] = selected_mappings
        request.session.pop(_tentative_key(project.id, "result"), None)
        messages.info(request, "Tentative BOM uploaded. Review the column mapping, then validate.")
        return redirect(f"{redirect('estimation:estimate_detail', project_id=project.id).url}?sheet=raw-material-selection")

    tmp_path = request.session.get(_tentative_key(project.id, "tmp_path"))
    headers_info = request.session.get(_tentative_key(project.id, "headers_info")) or {}
    if not tmp_path or not headers_info:
        messages.error(request, "Please upload the tentative BOM file first.")
        return redirect(f"{redirect('estimation:estimate_detail', project_id=project.id).url}?sheet=raw-material-selection")
    if not os.path.exists(tmp_path):
        _clear_tentative_bom_session(request, project.id)
        messages.error(request, "The uploaded tentative BOM file is no longer available on the server. Please upload it again.")
        return redirect(f"{redirect('estimation:estimate_detail', project_id=project.id).url}?sheet=raw-material-selection")

    posted_mappings = _build_tentative_user_sheet_mappings(request, headers_info)
    selected_mappings = request.session.get(_tentative_key(project.id, "selected_mappings")) or {}
    if _has_any_mapping(posted_mappings):
        selected_mappings = posted_mappings
        request.session[_tentative_key(project.id, "selected_mappings")] = selected_mappings

    if _has_any_mapping(selected_mappings):
        _persist_tentative_mappings(headers_info, selected_mappings, request.user)

    result = validate_and_extract_tentative_bom(tmp_path, user_sheet_mappings=selected_mappings)
    total_quantity_mt = sum((row["quantity_mt"] for row in result.get("aggregated_lines", [])), Decimal("0"))
    request.session[_tentative_key(project.id, "result")] = {
        "ok": result["ok"],
        "errors": _session_safe_errors(result.get("errors", [])),
        "sheets_used": result.get("sheets_used", 0),
        "matched_rows": result.get("matched_rows", 0),
        "total_quantity_mt": str(total_quantity_mt.quantize(Decimal("0.001"))),
        "aggregated_lines": [
            {
                "item_description": row["item"].item_description,
                "grade_name": row["item"].grade.name,
                "section_name": row["item"].section_name,
                "quantity_mt": str(row["quantity_mt"]),
            }
            for row in result.get("aggregated_lines", [])
        ],
    }

    if action == "import":
        if not result["ok"]:
            messages.error(request, "Tentative BOM has validation errors. Fix them before import.")
            return redirect(f"{redirect('estimation:estimate_detail', project_id=project.id).url}?sheet=raw-material-selection")

        project.raw_material_lines.all().delete()
        created_count = 0
        for index, row in enumerate(result["aggregated_lines"], start=1):
            EstimateRawMaterialLine.objects.create(
                project=project,
                item=row["item"],
                quantity_mt=row["quantity_mt"],
                sort_order=index,
            )
            created_count += 1
        sync_project_supplier_rates(project)
        recalculate_cost_heads(project)
        project.status = EstimateProject.Status.RATE_FINALIZATION
        project.updated_by = request.user
        project.save(update_fields=["status", "updated_by", "updated_at"])
        _clear_tentative_bom_session(request, project.id)
        messages.success(request, f"Tentative BOM imported and {created_count} raw material line(s) created.")
        return redirect(f"{redirect('estimation:estimate_detail', project_id=project.id).url}?sheet=raw-material-selection")

    if result["ok"]:
        messages.success(request, "Tentative BOM validation passed. Import is now enabled.")
    else:
        messages.error(request, "Tentative BOM validation found errors. Please correct the mapping or source file.")
    return redirect(f"{redirect('estimation:estimate_detail', project_id=project.id).url}?sheet=raw-material-selection")


@login_required
def add_raw_material_line(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_raw_materials(request.user):
        messages.error(request, "Only Planning, Management, or Admin can select raw materials.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    item_id = request.POST.get("item_id")
    quantity_mt = _parse_decimal(request.POST.get("quantity_mt"))
    if not item_id or quantity_mt <= 0:
        messages.error(request, "Select a raw material item and enter quantity in MT.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    item = get_object_or_404(Item, pk=item_id, is_active=True)
    EstimateRawMaterialLine.objects.create(
        project=project,
        item=item,
        quantity_mt=quantity_mt,
        sort_order=project.raw_material_lines.count() + 1,
    )
    sync_project_supplier_rates(project)
    recalculate_cost_heads(project)
    project.status = EstimateProject.Status.RATE_FINALIZATION
    project.updated_by = request.user
    project.save(update_fields=["status", "updated_by", "updated_at"])
    messages.success(request, "Raw material line added.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def delete_raw_material_line(request, project_id: int, line_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_raw_materials(request.user):
        messages.error(request, "You do not have permission to delete raw material lines.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    line = get_object_or_404(EstimateRawMaterialLine, pk=line_id, project=project)
    line.delete()

    for index, row in enumerate(project.raw_material_lines.order_by("sort_order", "id"), start=1):
        if row.sort_order != index:
            row.sort_order = index
            row.save(update_fields=["sort_order"])

    recalculate_cost_heads(project)
    project.updated_by = request.user
    project.save(update_fields=["updated_by", "updated_at"])
    messages.success(request, "Raw material line deleted.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def update_rate_sheet(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_rates(request.user):
        messages.error(request, "You do not have permission to update supplier rates.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    for line in project.raw_material_lines.prefetch_related("supplier_rates"):
        final_key = f"final_rate_{line.id}"
        line.final_rate_per_mt = _parse_decimal(request.POST.get(final_key))
        line.save(update_fields=["final_rate_per_mt"])
        for rate in line.supplier_rates.all():
            field_name = f"rate_{line.id}_{rate.supplier_id}"
            rate.rate_per_mt = _parse_decimal(request.POST.get(field_name), default=None)
            rate.save(update_fields=["rate_per_mt"])
        line.recalculate_from_rates(save=True)

    recalculate_cost_heads(project)
    if project.status in {
        EstimateProject.Status.DRAFT,
        EstimateProject.Status.RATE_FINALIZATION,
        EstimateProject.Status.REJECTED,
        EstimateProject.Status.UNDER_REVIEW,
    }:
        project.status = EstimateProject.Status.UNDER_REVIEW
    project.updated_by = request.user
    update_fields = ["updated_by", "updated_at"]
    if project.status == EstimateProject.Status.UNDER_REVIEW:
        update_fields.insert(0, "status")
    project.save(update_fields=update_fields)
    messages.success(request, "Rate finalization sheet updated.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def download_rate_sheet(request, project_id: int):
    project = get_object_or_404(
        EstimateProject.objects.prefetch_related(
            "project_suppliers__supplier",
            "raw_material_lines__item__grade",
            "raw_material_lines__supplier_rates__supplier",
        ),
        pk=project_id,
    )
    supplier_links = list(project.project_suppliers.select_related("supplier"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rate Finalisation"
    ws["A1"] = "KALPADEEP INDUSTRIES PVT LTD"
    ws["A2"] = f"Rate Sheet - {project.inquiry_no}"
    ws["A3"] = f"Client: {project.client_name}"
    ws["D3"] = f"Project: {project.project_name}"
    ws["A5"] = "Line ID"
    ws["B5"] = "Item Description"
    ws["C5"] = "Grade"
    ws["D5"] = "Section"
    ws["E5"] = "Quantity MT"

    supplier_start_col = 6
    for idx, link in enumerate(supplier_links, start=supplier_start_col):
        ws.cell(row=5, column=idx, value=f"{link.supplier.name} Rate/MT")

    lowest_rate_col = supplier_start_col + len(supplier_links)
    final_rate_col = lowest_rate_col + 1
    ws.cell(row=5, column=lowest_rate_col, value="Lowest (L1) Rate/MT")
    ws.cell(row=5, column=final_rate_col, value="Final Rate/MT")
    ws.cell(row=5, column=final_rate_col + 1, value="Total Amount")

    current_row = 6
    for line in project.raw_material_lines.all():
        supplier_map = {rate.supplier_id: rate for rate in line.supplier_rates.all()}
        ws.cell(row=current_row, column=1, value=line.id)
        ws.cell(row=current_row, column=2, value=line.item.item_description)
        ws.cell(row=current_row, column=3, value=line.item.grade.name)
        ws.cell(row=current_row, column=4, value=line.item.section_name)
        ws.cell(row=current_row, column=5, value=float(line.quantity_mt or 0))
        for idx, link in enumerate(supplier_links, start=supplier_start_col):
            rate = supplier_map.get(link.supplier_id)
            ws.cell(row=current_row, column=idx, value=float(rate.rate_per_mt) if rate and rate.rate_per_mt is not None else "")
        ws.cell(row=current_row, column=lowest_rate_col, value=float(line.lowest_rate_per_mt or 0))
        ws.cell(row=current_row, column=final_rate_col, value=float(line.final_rate_per_mt or 0))
        ws.cell(row=current_row, column=final_rate_col + 1, value=float(line.total_amount or 0))
        current_row += 1

    ws.column_dimensions["A"].hidden = True
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{project.inquiry_no}_rate_sheet.xlsx"'
    return response


@login_required
def upload_rate_sheet(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_rates(request.user):
        messages.error(request, "Only Marketing, Management, or Admin can upload the supplier rate sheet.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please choose the filled supplier rate sheet to upload.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    wb = openpyxl.load_workbook(upload, data_only=True)
    ws = wb.active
    header_row = None
    for row_no in range(1, min(ws.max_row, 15) + 1):
        row_values = [str(ws.cell(row=row_no, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
        if "Line ID" in row_values and "Item Description" in row_values:
            header_row = row_no
            break

    if not header_row:
        messages.error(request, "Uploaded file does not look like the system-generated rate sheet.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    headers = {
        str(ws.cell(row=header_row, column=col).value or "").strip(): col
        for col in range(1, ws.max_column + 1)
    }
    line_id_col = headers.get("Line ID")
    final_rate_col = headers.get("Final Rate/MT")
    supplier_columns = {}
    for link in project.project_suppliers.select_related("supplier"):
        header = f"{link.supplier.name} Rate/MT"
        if header in headers:
            supplier_columns[link.supplier_id] = headers[header]

    updated_lines = 0
    for row_no in range(header_row + 1, ws.max_row + 1):
        raw_line_id = ws.cell(row=row_no, column=line_id_col).value if line_id_col else None
        if raw_line_id in (None, ""):
            continue
        try:
            line_id = int(raw_line_id)
        except (TypeError, ValueError):
            continue
        try:
            line = project.raw_material_lines.prefetch_related("supplier_rates").get(pk=line_id)
        except EstimateRawMaterialLine.DoesNotExist:
            continue

        supplier_rate_map = {rate.supplier_id: rate for rate in line.supplier_rates.all()}
        for supplier_id, col_no in supplier_columns.items():
            rate = supplier_rate_map.get(supplier_id)
            if not rate:
                continue
            cell_value = ws.cell(row=row_no, column=col_no).value
            parsed = _parse_decimal("" if cell_value is None else str(cell_value), default=None)
            rate.rate_per_mt = parsed
            rate.save(update_fields=["rate_per_mt"])

        final_value = ws.cell(row=row_no, column=final_rate_col).value if final_rate_col else None
        parsed_final = _parse_decimal("" if final_value is None else str(final_value), default=Decimal("0"))
        line.final_rate_per_mt = parsed_final
        line.save(update_fields=["final_rate_per_mt"])
        line.recalculate_from_rates(save=True)
        updated_lines += 1

    recalculate_cost_heads(project)
    if project.status in {
        EstimateProject.Status.DRAFT,
        EstimateProject.Status.RATE_FINALIZATION,
        EstimateProject.Status.REJECTED,
        EstimateProject.Status.UNDER_REVIEW,
    }:
        project.status = EstimateProject.Status.UNDER_REVIEW
    project.updated_by = request.user
    project.save(update_fields=["status", "updated_by", "updated_at"])
    messages.success(request, f"Supplier rate sheet uploaded and {updated_lines} raw material line(s) refreshed.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def upload_supplier_quotation(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_rates(request.user):
        messages.error(request, "Only Marketing, Management, or Admin can upload supplier quotation files.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    upload = request.FILES.get("quotation_file")
    supplier_id = request.POST.get("supplier_id")
    remarks = (request.POST.get("remarks") or "").strip()
    if not upload or not supplier_id:
        messages.error(request, "Supplier and quotation file are required.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    supplier = get_object_or_404(EstimateSupplier, pk=supplier_id)
    object_key = build_supplier_quotation_object_key(project, supplier.name, upload.name)
    upload_supplier_quotation_file(
        upload,
        object_key,
        content_type=getattr(upload, "content_type", "") or "application/octet-stream",
    )
    EstimateSupplierQuotationFile.objects.create(
        project=project,
        supplier=supplier,
        file_key=object_key,
        original_filename=upload.name,
        content_type=getattr(upload, "content_type", "") or "",
        file_size=getattr(upload, "size", None),
        remarks=remarks,
        uploaded_by=request.user,
    )
    messages.success(request, f"Supplier quotation file uploaded for {supplier.name}.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def update_cost_heads(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_costs(request.user):
        messages.error(request, "You do not have permission to update estimation cost heads.")
        return redirect("estimation:estimate_detail", project_id=project.id)
    if project.status == EstimateProject.Status.CLOSED:
        messages.error(request, "Closed budgets cannot be modified.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    for head in project.cost_heads.filter(line_type="ENTRY"):
        if head.is_percentage_editable:
            entered_percentage = _parse_decimal(
                request.POST.get(
                    f"consumption_{head.id}" if head.code in PAINT_COMPONENT_CODES else f"percentage_{head.id}"
                )
            )
            head.percentage = (
                entered_percentage
                if head.code in PAINT_COMPONENT_CODES
                else entered_percentage / Decimal("100")
            )
        if head.is_rate_editable and head.code != "RAW_MATERIAL_COST":
            head.rate_per_kg = _parse_decimal(request.POST.get(f"rate_{head.id}"))
        head.remarks = (request.POST.get(f"remarks_{head.id}") or "").strip()
        head.save(update_fields=["percentage", "rate_per_kg", "remarks"])

    recalculate_cost_heads(project)
    if project.status in {
        EstimateProject.Status.DRAFT,
        EstimateProject.Status.RATE_FINALIZATION,
        EstimateProject.Status.REJECTED,
        EstimateProject.Status.UNDER_REVIEW,
    }:
        project.status = EstimateProject.Status.UNDER_REVIEW
    project.updated_by = request.user
    update_fields = ["updated_by", "updated_at"]
    if project.status == EstimateProject.Status.UNDER_REVIEW:
        update_fields.insert(0, "status")
    project.save(update_fields=update_fields)
    messages.success(request, "Quotation calculation updated.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def update_department_notes(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if project.status == EstimateProject.Status.CLOSED:
        messages.error(request, "Closed budgets cannot be modified.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    update_fields = ["updated_by", "updated_at"]
    updated = False
    role = request.user.role

    if role in {"Admin", "Planning", "Management"}:
        project.planning_notes = (request.POST.get("planning_notes") or "").strip()
        update_fields.append("planning_notes")
        updated = True
    if role in {"Admin", "Marketing", "Management"}:
        project.marketing_notes = (request.POST.get("marketing_notes") or "").strip()
        update_fields.append("marketing_notes")
        updated = True
    if role in {"Admin", "Management"}:
        project.management_notes = (request.POST.get("management_notes") or "").strip()
        update_fields.append("management_notes")
        updated = True
    if role in {"Admin", "Accounts", "Management"}:
        project.accounts_notes = (request.POST.get("accounts_notes") or "").strip()
        update_fields.append("accounts_notes")
        updated = True

    if not updated:
        messages.error(request, "You do not have permission to update department notes.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    project.updated_by = request.user
    project.save(update_fields=update_fields)
    messages.success(request, "Department notes updated.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def submit_management_decision(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_decision(request.user):
        messages.error(request, "Only Management or Admin can approve or reject quotations.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    action = (request.POST.get("decision") or "").strip().lower()
    quoted_price_per_mt = _parse_decimal(request.POST.get("quoted_price_per_mt"))
    approved_price_per_mt = _parse_decimal(request.POST.get("approved_price_per_mt"))
    decision_notes = (request.POST.get("decision_notes") or "").strip()
    management_notes = (request.POST.get("management_notes") or "").strip()

    if action not in {"approve", "reject"}:
        messages.error(request, "Choose whether the quotation is approved or rejected.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    project.quoted_price_per_mt = quoted_price_per_mt
    project.approved_price_per_mt = approved_price_per_mt
    project.decision_notes = decision_notes
    project.management_notes = management_notes
    project.decision_by = request.user
    project.decision_at = timezone.now()
    project.updated_by = request.user
    project.status = (
        EstimateProject.Status.APPROVED if action == "approve" else EstimateProject.Status.REJECTED
    )
    project.save(
        update_fields=[
            "quoted_price_per_mt",
            "approved_price_per_mt",
            "decision_notes",
            "management_notes",
            "decision_by",
            "decision_at",
            "status",
            "updated_by",
            "updated_at",
        ]
    )
    messages.success(
        request,
        "Quotation approved and saved." if action == "approve" else "Quotation rejected and saved.",
    )
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def reopen_quotation_for_review(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_decision(request.user):
        messages.error(request, "Only Management or Admin can reopen quotations for review.")
        return redirect("estimation:estimate_detail", project_id=project.id)
    if project.status not in {EstimateProject.Status.APPROVED, EstimateProject.Status.REJECTED}:
        messages.error(request, "Only approved or rejected quotations can be reopened.")
        return redirect("estimation:estimate_detail", project_id=project.id)
    if project.budget_heads.exists():
        messages.error(request, "This quotation cannot be reopened because the budget sheet has already been created.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    project.status = EstimateProject.Status.UNDER_REVIEW
    project.updated_by = request.user
    project.save(update_fields=["status", "updated_by", "updated_at"])
    messages.success(request, "Quotation reopened for review.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def mark_po_received(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_accounts(request.user):
        messages.error(request, "You do not have permission to mark PO received.")
        return redirect("estimation:estimate_detail", project_id=project.id)
    if project.status == EstimateProject.Status.CLOSED:
        messages.error(request, "Closed budgets cannot be modified.")
        return redirect("estimation:estimate_detail", project_id=project.id)
    if project.status != EstimateProject.Status.APPROVED:
        messages.error(request, "Only an approved quotation can move to PO and budget monitoring.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    project.work_order_no = (request.POST.get("work_order_no") or "").strip()
    project.purchase_order_no = (request.POST.get("purchase_order_no") or "").strip()
    project.purchase_order_date = _parse_date(request.POST.get("purchase_order_date"))
    project.delivery_date = _parse_date(request.POST.get("delivery_date"))
    project.status = EstimateProject.Status.PO_RECEIVED
    project.updated_by = request.user
    project.save()
    generate_budget_heads(project)
    refresh_budget_totals(project)
    messages.success(request, "PO details saved and budget heads generated.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def add_expense(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_accounts(request.user):
        messages.error(request, "You do not have permission to add expenditures.")
        return redirect("estimation:estimate_detail", project_id=project.id)
    if project.status == EstimateProject.Status.CLOSED:
        messages.error(request, "Closed budgets cannot accept new expenditures.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    budget = get_object_or_404(project.budget_heads.all(), pk=request.POST.get("budget_head_id"))
    amount = _parse_decimal(request.POST.get("amount"))
    description = (request.POST.get("description") or "").strip()
    if amount <= 0 or not description:
        messages.error(request, "Budget head, amount, and description are required.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    EstimateExpense.objects.create(
        budget_head=budget,
        expense_date=request.POST.get("expense_date") or timezone.now().date(),
        reference_no=(request.POST.get("reference_no") or "").strip(),
        amount=amount,
        description=description,
        created_by=request.user,
    )
    refresh_budget_totals(project)
    messages.success(request, "Expenditure recorded and marked pending for approval.")
    return redirect("estimation:estimate_detail", project_id=project.id)


@login_required
def approve_expense(request, expense_id: int):
    expense = get_object_or_404(EstimateExpense.objects.select_related("budget_head__project"), pk=expense_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=expense.budget_head.project_id)
    if request.user.role not in {"Admin", "Management"}:
        messages.error(request, "Only management can approve expenditures.")
        return redirect("estimation:estimate_detail", project_id=expense.budget_head.project_id)
    if expense.budget_head.project.status == EstimateProject.Status.CLOSED:
        messages.error(request, "Closed budgets cannot be modified.")
        return redirect("estimation:estimate_detail", project_id=expense.budget_head.project_id)

    expense.status = EstimateExpense.Status.APPROVED
    expense.approved_by = request.user
    expense.approved_at = timezone.now()
    expense.save(update_fields=["status", "approved_by", "approved_at"])
    refresh_budget_totals(expense.budget_head.project)
    messages.success(request, "Expenditure approved.")
    return redirect("estimation:estimate_detail", project_id=expense.budget_head.project_id)


@login_required
def close_budget(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not _can_manage_accounts(request.user):
        messages.error(request, "Only Accounts, Management, or Admin can close a budget.")
        return redirect("estimation:estimate_detail", project_id=project.id)
    if not project.work_order_no:
        messages.error(request, "A work order number is required before closing a budget.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    project.status = EstimateProject.Status.CLOSED
    project.updated_by = request.user
    project.save(update_fields=["status", "updated_by", "updated_at"])
    messages.success(request, "Budget closed and moved to closed budgets.")
    return redirect("estimation:estimate_list")


@login_required
def delete_estimate(request, project_id: int):
    project = get_object_or_404(EstimateProject, pk=project_id)
    if request.method != "POST":
        return redirect("estimation:estimate_list")
    if not _can_delete_estimate(request.user):
        messages.error(request, "Only Management or Admin can delete quotations.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    if project.boms.exists():
        messages.error(request, "This quotation cannot be deleted because it is already linked with BOM records.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    if project.budget_heads.exists() or EstimateExpense.objects.filter(budget_head__project=project).exists():
        messages.error(request, "This quotation cannot be deleted because budget or expenditure records already exist.")
        return redirect("estimation:estimate_detail", project_id=project.id)

    project_name = project.project_name
    project.delete()
    messages.success(request, f'Quotation "{project_name}" deleted successfully.')
    return redirect("estimation:estimate_list")


@login_required
def export_quotation_excel(request, project_id: int):
    project = get_object_or_404(EstimateProject.objects.prefetch_related("cost_heads"), pk=project_id)
    recalculate_cost_heads(project)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    ws["A1"] = "KALPADEEP INDUSTRIES PVT LTD"
    ws["A2"] = "Quotation Sheet"
    ws["A3"] = f"Client: {project.client_name}"
    ws["C3"] = f"Project: {project.project_name}"
    ws["E3"] = f"Quantity MT: {float(project.quantity_mt)}"
    ws["A4"] = f"Status: {project.get_status_display()}"
    ws["C4"] = f"Estimated Price /MT: {float(project.estimated_price_per_mt or 0)}"
    ws["E4"] = f"Quoted Price /MT: {float(project.quoted_price_per_mt or 0)}"
    ws["G4"] = f"Approved Price /MT: {float(project.approved_price_per_mt or 0)}"
    ws.append([])
    ws.append(["Cost Head", "Percentage / Consumption", "Cost per Kg / Rate per LTR", "Cost", "Remarks"])
    for head in project.cost_heads.all():
        second_column = (
            float(head.percentage or 0)
            if head.code in PAINT_COMPONENT_CODES
            else float((head.percentage or 0) * Decimal("100")) if head.percentage is not None else ""
        )
        ws.append([
            head.name,
            second_column,
            float(head.rate_per_kg or 0),
            float(head.amount or 0),
            head.remarks,
        ])

    raw_ws = wb.create_sheet("Raw Material Selection")
    raw_ws.append(["Item Description", "Grade", "Section", "Quantity MT", "Final Rate/MT", "Total Amount"])
    for line in project.raw_material_lines.select_related("item__grade").all():
        raw_ws.append([
            line.item.item_description,
            line.item.grade.name,
            line.item.section_name,
            float(line.quantity_mt or 0),
            float(line.final_rate_per_mt or 0),
            float(line.total_amount or 0),
        ])

    rate_ws = wb.create_sheet("Rate Finalisation")
    supplier_links = list(project.project_suppliers.select_related("supplier"))
    rate_headers = ["Item Description", "Grade", "Section", "Quantity MT"] + [
        f"{link.supplier.name} Rate/MT" for link in supplier_links
    ] + ["Lowest (L1) Rate/MT", "Final Rate/MT", "Total Amount"]
    rate_ws.append(rate_headers)
    for line in project.raw_material_lines.select_related("item__grade").prefetch_related("supplier_rates").all():
        supplier_map = {rate.supplier_id: rate for rate in line.supplier_rates.all()}
        row = [
            line.item.item_description,
            line.item.grade.name,
            line.item.section_name,
            float(line.quantity_mt or 0),
        ]
        for link in supplier_links:
            row.append(float((supplier_map.get(link.supplier_id).rate_per_mt if supplier_map.get(link.supplier_id) else 0) or 0))
        row.extend([float(line.lowest_rate_per_mt or 0), float(line.final_rate_per_mt or 0), float(line.total_amount or 0)])
        rate_ws.append(row)

    budget_ws = wb.create_sheet("Budget Monitoring")
    budget_ws.append(["Budget Code", "Budget Head", "Budget Amount", "Spent Amount", "Approved Amount"])
    for budget in project.budget_heads.all():
        budget_ws.append([
            budget.budget_code,
            budget.name,
            float(budget.budget_amount or 0),
            float(budget.spent_amount or 0),
            float(budget.approved_amount or 0),
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{project.inquiry_no}_quotation.xlsx"'
    return response


@login_required
def export_quotation_pdf(request, project_id: int):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except Exception:
        messages.error(request, "PDF export package is not installed on this server yet.")
        return redirect("estimation:estimate_detail", project_id=project_id)

    project = get_object_or_404(EstimateProject.objects.prefetch_related("cost_heads"), pk=project_id)
    recalculate_cost_heads(project)

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 20 * mm

    logo_path = Path(settings.BASE_DIR) / "static" / "logo" / "kalpadeeplogo.png"
    if logo_path.exists():
        try:
            pdf.drawImage(str(logo_path), 15 * mm, height - 35 * mm, width=25 * mm, height=18 * mm, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(45 * mm, y, "KALPADEEP INDUSTRIES PVT LTD")
    y -= 8 * mm
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(15 * mm, y, "Quotation")
    y -= 7 * mm
    pdf.setFont("Helvetica", 10)
    pdf.drawString(15 * mm, y, f"Client: {project.client_name}")
    y -= 6 * mm
    pdf.drawString(15 * mm, y, f"Project: {project.project_name}")
    y -= 6 * mm
    pdf.drawString(15 * mm, y, f"Quantity (MT): {project.quantity_mt}")
    y -= 6 * mm
    pdf.drawString(15 * mm, y, f"Status: {project.get_status_display()}")
    y -= 6 * mm
    pdf.drawString(15 * mm, y, f"Estimated Price /MT: {project.estimated_price_per_mt}")
    y -= 6 * mm
    pdf.drawString(15 * mm, y, f"Quoted Price /MT: {project.quoted_price_per_mt}")
    y -= 6 * mm
    pdf.drawString(15 * mm, y, f"Approved Price /MT: {project.approved_price_per_mt}")
    y -= 10 * mm

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(15 * mm, y, "Cost Head")
    pdf.drawString(78 * mm, y, "Pct / Cons.")
    pdf.drawString(114 * mm, y, "Cost/Kg / Rate")
    pdf.drawString(145 * mm, y, "Cost")
    y -= 5 * mm
    pdf.line(15 * mm, y, 195 * mm, y)
    y -= 5 * mm
    pdf.setFont("Helvetica", 9)

    for head in project.cost_heads.all():
        if y < 25 * mm:
            pdf.showPage()
            y = height - 20 * mm
            pdf.setFont("Helvetica", 9)
        pct = ""
        if head.percentage is not None:
            if head.code in PAINT_COMPONENT_CODES:
                pct = f"{head.percentage.quantize(Decimal('0.001'))} LTR/MT"
            else:
                pct_value = head.percentage * Decimal("100")
                pct = f"{pct_value.quantize(Decimal('0.001'))}%"
        pdf.drawString(15 * mm, y, str(head.name)[:42])
        pdf.drawRightString(108 * mm, y, pct)
        pdf.drawRightString(140 * mm, y, f"{head.rate_per_kg}")
        pdf.drawRightString(188 * mm, y, f"{head.amount}")
        y -= 5.5 * mm

    y -= 10 * mm
    pdf.line(120 * mm, y, 190 * mm, y)
    y -= 5 * mm
    pdf.drawString(125 * mm, y, "Quotation Approved By")
    y -= 12 * mm
    pdf.line(120 * mm, y, 190 * mm, y)
    y -= 5 * mm
    pdf.drawString(145 * mm, y, "Signature")

    pdf.save()
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{project.inquiry_no}_quotation.pdf"'
    return response
