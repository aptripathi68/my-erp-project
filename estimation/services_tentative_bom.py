from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
import hashlib
import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from masters.models import Item, normalize_item_description
from procurement.services.bom_importer import normalize_grade_name

DEFAULT_GRADE_CODE = "IS:2062"
DEFAULT_GRADE_NAME = "E250BR"
DEFAULT_GRADE_DISPLAY = f"{DEFAULT_GRADE_CODE} {DEFAULT_GRADE_NAME}"


def _h(value: Any) -> str:
    if value is None:
        return ""
    value = str(value).strip().lower().replace("\n", " ")
    return re.sub(r"\s+", " ", value)


def _to_decimal(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except Exception:
            return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


ALIASES = {
    "section_name": [
        "thickness",
        "section name",
        "section",
        "section size",
        "section designation",
        "member section",
        "profile size",
        "size",
        "material",
        "item description",
        "profile",
        "member",
        "material description",
    ],
    "grade": [
        "grade",
        "grade name",
        "material grade",
        "mat grade",
        "steel grade",
        "material spec",
        "specification",
        "spec",
        "quality",
    ],
    "gross_weight": [
        "drg gross wt.",
        "drg gross wt",
        "drg wt.",
        "drg wt",
        "gross weight",
        "gross wt",
        "gross wt (kg)",
        "gross weight (kg)",
        "drawing gross weight",
        "drawing gross wt",
        "drawing gross wt (kg)",
        "engg weight",
        "engg weight (kg)",
        "unit wt",
        "unit weight",
        "weight (kg)",
        "total weight",
        "weight",
        "wt",
    ],
}

IGNORE_SHEET_NAME_CONTAINS = ["summary", "notes", "index", "cover"]


@dataclass
class TentativeBOMRow:
    sheet_name: str
    excel_row: int
    section_name_raw: str
    grade_raw: str
    gross_weight_kg: Decimal
    item_id: int


def detect_header_row(ws, max_scan_rows: int = 40):
    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        headers = [_h(v) for v in row]
        hits = 0
        for field_aliases in ALIASES.values():
            if any(h in field_aliases for h in headers):
                hits += 1
        if hits >= 2:
            return row_index, headers
    return None, None


def build_col_map(headers: List[str]) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for canonical, aliases in ALIASES.items():
        for alias in aliases:
            for idx, header in enumerate(headers):
                if header == alias:
                    col_map[canonical] = idx
                    break
            if canonical in col_map:
                break
    return col_map


def build_user_col_map(headers: List[str], user_mapping: Dict[str, str]) -> Dict[str, int]:
    normalized_headers = [_h(h) for h in headers]
    col_map: Dict[str, int] = {}
    for canonical, selected_header in user_mapping.items():
        selected_norm = _h(selected_header)
        if not selected_norm:
            continue
        for idx, header_norm in enumerate(normalized_headers):
            if header_norm == selected_norm:
                col_map[canonical] = idx
                break
    return col_map


def build_header_signature(headers: List[str]) -> str:
    normalized = [_h(h) for h in headers if _h(h)]
    return hashlib.sha256("|".join(normalized).encode("utf-8")).hexdigest()


def get_cell(row: Tuple[Any], col_map: Dict[str, int], key: str):
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def workbook_sheet_headers(xlsx_path: str, max_scan_rows: int = 40) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    result: Dict[str, Any] = {}
    for sheet_name in wb.sheetnames:
        if any(x in sheet_name.lower() for x in IGNORE_SHEET_NAME_CONTAINS):
            continue
        ws = wb[sheet_name]
        header_row, headers = detect_header_row(ws, max_scan_rows=max_scan_rows)
        if not header_row or not headers:
            result[sheet_name] = {"detected": False, "header_row": None, "headers": []}
            continue
        auto_col_map = build_col_map(headers)
        auto_mapping = {k: headers[idx] for k, idx in auto_col_map.items() if idx < len(headers)}
        result[sheet_name] = {
            "detected": True,
            "header_row": header_row,
            "headers": headers,
            "header_signature": build_header_signature(headers),
            "mapping": auto_mapping,
        }
    return result


def validate_and_extract_tentative_bom(
    xlsx_path: str,
    user_sheet_mappings: Optional[Dict[str, Dict[str, str]]] = None,
) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    items = list(
        Item.objects.select_related("grade")
        .filter(is_active=True)
        .only("id", "section_name", "item_description", "grade__code", "grade__name")
    )

    item_by_section_grade: Dict[Tuple[str, str], Item] = {}
    grade_candidates_by_section: Dict[str, List[Tuple[str, Item]]] = {}
    for item in items:
        section_norm = normalize_item_description(item.section_name or "")
        if not section_norm or not item.grade:
            continue
        grade_code_norm = normalize_grade_name(item.grade.code or "")
        grade_name_norm = normalize_grade_name(item.grade.name or "")
        grade_combined_norm = normalize_grade_name(f"{item.grade.code or ''} {item.grade.name or ''}")
        for grade_norm in (grade_code_norm, grade_name_norm, grade_combined_norm):
            if grade_norm:
                item_by_section_grade[(section_norm, grade_norm)] = item
                grade_candidates_by_section.setdefault(section_norm, []).append((grade_norm, item))

    extracted: List[TentativeBOMRow] = []
    errors: List[Dict[str, Any]] = []
    detected: Dict[str, Any] = {}
    sheets_used = 0

    for sheet_name in wb.sheetnames:
        if any(x in sheet_name.lower() for x in IGNORE_SHEET_NAME_CONTAINS):
            continue
        ws = wb[sheet_name]
        header_row, headers = detect_header_row(ws)
        if not header_row or not headers:
            detected[sheet_name] = {"skipped": True, "reason": "header_not_detected"}
            continue

        sheet_mapping = (user_sheet_mappings or {}).get(sheet_name, {})
        col_map = build_user_col_map(headers, sheet_mapping) if sheet_mapping else build_col_map(headers)

        missing_fields = [field for field in ("section_name", "gross_weight") if field not in col_map]
        if missing_fields:
            detected[sheet_name] = {"skipped": True, "reason": f"missing_columns:{','.join(missing_fields)}"}
            continue

        sheets_used += 1
        detected[sheet_name] = {"skipped": False, "header_row": header_row, "col_map": col_map}

        for excel_row, row_vals in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not row_vals or all(v is None or str(v).strip() == "" for v in row_vals):
                continue

            section_val = get_cell(row_vals, col_map, "section_name")
            grade_val = get_cell(row_vals, col_map, "grade")
            gross_weight_val = get_cell(row_vals, col_map, "gross_weight")

            section_raw = str(section_val).strip() if section_val is not None else ""
            grade_raw = str(grade_val).strip() if grade_val is not None else ""
            gross_weight = _to_decimal(gross_weight_val)

            if not section_raw and not grade_raw and gross_weight is None:
                continue

            # Ignore section/group headers and assembly summary rows that do not identify a raw material profile.
            if not section_raw:
                continue

            row_errors: List[str] = []
            if not grade_raw:
                grade_raw = DEFAULT_GRADE_DISPLAY
            if gross_weight is None or gross_weight <= 0:
                continue

            matched_item = None
            if section_raw and grade_raw:
                section_norm = normalize_item_description(section_raw)
                grade_norm = normalize_grade_name(grade_raw)
                matched_item = item_by_section_grade.get((section_norm, grade_norm))
                if matched_item is None:
                    for candidate_grade, candidate_item in grade_candidates_by_section.get(section_norm, []):
                        if grade_norm and (grade_norm in candidate_grade or candidate_grade in grade_norm):
                            matched_item = candidate_item
                            break
                if matched_item is None:
                    row_errors.append("No Item Master match for Section Name + Grade.")

            if row_errors:
                errors.append(
                    {
                        "sheet_name": sheet_name,
                        "excel_row": excel_row,
                        "section_name": section_raw,
                        "grade": grade_raw,
                        "gross_weight": gross_weight_val,
                        "errors": row_errors,
                    }
                )
                continue

            extracted.append(
                TentativeBOMRow(
                    sheet_name=sheet_name,
                    excel_row=excel_row,
                    section_name_raw=section_raw,
                    grade_raw=grade_raw,
                    gross_weight_kg=gross_weight,
                    item_id=matched_item.id,
                )
            )

    aggregated: Dict[int, Dict[str, Any]] = {}
    for row in extracted:
        bucket = aggregated.setdefault(
            row.item_id,
            {
                "item": Item.objects.select_related("grade").get(pk=row.item_id),
                "gross_weight_kg": Decimal("0"),
                "source_rows": 0,
            },
        )
        bucket["gross_weight_kg"] += row.gross_weight_kg
        bucket["source_rows"] += 1

    aggregated_lines = []
    for item_id, bucket in aggregated.items():
        gross_weight_kg = bucket["gross_weight_kg"]
        aggregated_lines.append(
            {
                "item": bucket["item"],
                "gross_weight_kg": gross_weight_kg.quantize(Decimal("0.001")),
                "quantity_mt": (gross_weight_kg / Decimal("1000")).quantize(Decimal("0.001")),
                "source_rows": bucket["source_rows"],
            }
        )
    aggregated_lines.sort(key=lambda row: (row["item"].section_name, row["item"].grade.name, row["item"].item_description))

    return {
        "ok": sheets_used > 0 and not errors and bool(aggregated_lines),
        "errors": errors,
        "detected": detected,
        "sheets_used": sheets_used,
        "matched_rows": len(extracted),
        "aggregated_lines": aggregated_lines,
    }
